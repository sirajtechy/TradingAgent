#!/usr/bin/env python3
"""
run_halal_orchestrator_backtest_2025.py
========================================
Halal universe **orchestrator** (CWAF) backtest — January 2025 → December 2025.

Uses the orchestrator agent which fuses both:
  • **Technical Agent** — 12 frameworks, 40+ indicators, 8 pattern detectors
  • **Fundamental Agent** — Piotroski, Altman-Z, Graham, Greenblatt, Lynch,
      growth profile, Shariah compliance

via CWAF (Confidence-Weighted Asymmetric Fusion) to produce a single
composite signal per ticker per month.

Output: **LONG trades only** (bullish signals).

Inputs
------
  • data/halal_universe/halal_master.json  — 1 236 Shariah-screened tickers
  • Polygon.io API key (.env)              — primary OHLCV data source
  • FMP / yfinance                         — fundamental data

Ticker selection
----------------
  Per sector: 4 mega/large + 3 mid + 3 small/micro/nano = 10
  11 sectors (N/A excluded).  Total ≈ 105 tickers.

Monthly windows
---------------
  signal_date = 1st trading day of each month  (Jan → Dec 2025)
  result_date = last calendar day of each month

Trade setup (per LONG signal)
-----------------------------
  • entry_date         = signal_date
  • exit_date_est      = signal_date + holding_days × 1.4 calendar days
  • entry_price        = Polygon close at signal_date
  • target_price       = entry + 2.0 × ATR_14
  • stop_loss          = entry − 1.5 × ATR_14
  • expected_profit_pct= (target − entry) / entry × 100
  • confidence_score   = orchestrator_score  (0–100, CWAF)
  • profit_probability = confidence-mapped  (40–85 %)
  • direction          = LONG only  (bearish + neutral = skipped)
  • trade_duration_days= ADX-adjusted estimate  (filter: > 2 days only)
  • pattern_recognition= all detected chart patterns + details
  • fundamental_summary= fund subscores, data quality, band

Exports
-------
  • JSON : backtest_output/halal_orchestrator_2025/<TKR>_halal_orch_2025.json
  • Master: backtest_output/halal_orchestrator_2025/master_halal_orch_2025.json
  • Excel : backtest_output/halal_orchestrator_2025/halal_orchestrator_backtest_2025.xlsx
              sheets: Summary | All Trades | Pattern Log | <sector> × 11

Usage
-----
  # Full run (all sectors, 10 tickers each):
  python scripts/backtests/run_halal_orchestrator_backtest_2025.py

  # Single sector:
  python scripts/backtests/run_halal_orchestrator_backtest_2025.py --sector "Energy"

  # Override tickers:
  python scripts/backtests/run_halal_orchestrator_backtest_2025.py --tickers AAPL,MSFT,NVDA

  # Resume (skip tickers already saved):
  python scripts/backtests/run_halal_orchestrator_backtest_2025.py --resume

  # Quick smoke-test (1 ticker, 2 months):
  python scripts/backtests/run_halal_orchestrator_backtest_2025.py --tickers AAPL --months 2
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ── project root → sys.path ──────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

# Load .env before importing agent code (picks up POLYGON_API_KEY, FMP_API_KEY etc.)
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass

from agents.orchestrator.config import OrchestratorSettings
from agents.orchestrator.fusion import fuse_signals
from agents.polygon_data import PolygonClient as _PolygonBarFetcher
from agents.technical.data_client import PolygonTechnicalClient
import paths


# ─────────────────────────────────────────────────────────────────────────────
# Paths & constants
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_DIR      = paths.HALAL_ORCH_2025
HALAL_MASTER    = paths.HALAL_MASTER
TOP5_PER_SECTOR = paths.TOP5_PER_SECTOR

# ATR multipliers for target and stop (LONG setup)
ATR_TARGET_MULT = 2.0
ATR_STOP_MULT   = 1.5

# Pattern characteristic holding durations (trading days) — (min, max)
_PATTERN_CHAR_DURATIONS: Dict[str, Tuple[int, int]] = {
    "Cup & Handle":              (15, 22),
    "Ascending Triangle":        (10, 18),
    "Double Bottom":             (12, 20),
    "Inverse Head & Shoulders":  (15, 22),
    "Bull Flag":                 (5, 10),
    "Descending Triangle":       (8, 15),
    "Double Top":                (5, 8),
    "Head & Shoulders":          (5, 8),
}

# January – December 2025  (signal_date → result_date)
MONTHS_2025: List[Tuple[date, date]] = [
    (date(2025,  1,  1), date(2025,  1, 31)),
    (date(2025,  2,  1), date(2025,  2, 28)),
    (date(2025,  3,  1), date(2025,  3, 31)),
    (date(2025,  4,  1), date(2025,  4, 30)),
    (date(2025,  5,  1), date(2025,  5, 31)),
    (date(2025,  6,  1), date(2025,  6, 30)),
    (date(2025,  7,  1), date(2025,  7, 31)),
    (date(2025,  8,  1), date(2025,  8, 31)),
    (date(2025,  9,  1), date(2025,  9, 30)),
    (date(2025, 10,  1), date(2025, 10, 31)),
    (date(2025, 11,  1), date(2025, 11, 30)),
    (date(2025, 12,  1), date(2025, 12, 31)),
]

# Forward prediction window — Q1 + Q2 2026
PREDICTION_MONTHS: List[Tuple[date, date]] = [
    (date(2026,  1,  1), date(2026,  1, 31)),
    (date(2026,  2,  1), date(2026,  2, 28)),
    (date(2026,  3,  1), date(2026,  3, 31)),
    (date(2026,  4,  1), date(2026,  4, 30)),
    (date(2026,  5,  1), date(2026,  5, 31)),
    (date(2026,  6,  1), date(2026,  6, 30)),
]

# Singleton Polygon client for end-of-month price checks
_price_client = PolygonTechnicalClient()
# Direct Polygon bar fetcher for optimal-exit daily-bar lookups
_bar_fetcher  = _PolygonBarFetcher()

# Orchestrator settings
_ORCH_SETTINGS = OrchestratorSettings()


# ─────────────────────────────────────────────────────────────────────────────
# Ticker universe loader
# ─────────────────────────────────────────────────────────────────────────────

_LARGE_TIERS = {"mega", "large"}
_MID_TIERS   = {"mid"}
_SMALL_TIERS = {"small", "micro", "nano"}


def _load_ticker_universe(per_sector: int = 10) -> Dict[str, List[Dict]]:
    """
    Load halal_master.json and return {sector: [stock_meta, ...]} with
    ``per_sector`` tickers per entry, balanced across cap tiers:

        4 mega/large  +  3 mid  +  3 small/micro/nano  = 10

    Sectors with fewer than ``per_sector`` stocks get all available tickers.
    N/A sector is excluded.
    """
    with open(HALAL_MASTER) as fh:
        master = json.load(fh)

    result: Dict[str, List[Dict]] = {}

    for sector, stocks in master.get("stocks_by_sector", {}).items():
        if sector == "N/A":
            continue

        large = [s for s in stocks if s.get("market_cap_tier") in _LARGE_TIERS]
        mid   = [s for s in stocks if s.get("market_cap_tier") in _MID_TIERS]
        small = [s for s in stocks if s.get("market_cap_tier") in _SMALL_TIERS]

        n_large = min(4, len(large))
        n_mid   = min(3, len(mid))
        n_small = per_sector - n_large - n_mid

        selected: List[Dict] = (
            large[:n_large]
            + mid[:n_mid]
            + small[:max(0, n_small)]
        )

        # Pad to per_sector if still short
        ticker_set = {s["ticker"] for s in selected}
        for s in large + mid + small:
            if len(selected) >= per_sector:
                break
            if s["ticker"] not in ticker_set:
                selected.append(s)
                ticker_set.add(s["ticker"])

        if selected:
            for s in selected:
                tier = s.get("market_cap_tier", "unknown")
                if tier in _LARGE_TIERS:
                    s["cap_display"] = "large"
                elif tier in _MID_TIERS:
                    s["cap_display"] = "mid"
                else:
                    s["cap_display"] = "small"

            result[sector] = selected

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Trade setup calculator  (LONG only)
# ─────────────────────────────────────────────────────────────────────────────

def _compute_long_trade_setup(
    tech_result: Optional[Dict[str, Any]],
    fund_result: Optional[Dict[str, Any]],
    fusion_signal: str,                     # "bullish" | "neutral" | "bearish"
    orchestrator_score: float,
    final_confidence: float,
    signal_date: date,
    result_date: date,
    patterns: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Dict[str, Any]]:
    """
    Build a LONG trade setup from orchestrator output.

    Returns None when:
      - signal is NOT bullish (LONG only filter)
      - holding_days_est ≤ 2  (trade-duration filter)
      - required fields are absent
    """
    # ── LONG only ─────────────────────────────────────────────────────────
    if fusion_signal != "bullish":
        return None

    # ── Entry price from tech result ──────────────────────────────────────
    entry_price: Optional[float] = None
    for src in (tech_result, fund_result):
        if src and "as_of_price" in src:
            entry_price = src["as_of_price"]["price"]
            break
    if not entry_price or entry_price <= 0:
        return None

    # ── ATR & ADX from tech key_indicators ────────────────────────────────
    key_ind: Dict[str, Any] = {}
    if tech_result:
        key_ind = tech_result.get("key_indicators", {}) or {}

    atr: float = key_ind.get("atr_14") or (entry_price * 0.02)   # fallback: 2 %
    adx: Optional[float] = key_ind.get("adx")
    rsi: Optional[float] = key_ind.get("rsi_14")

    # ── LONG target & stop ────────────────────────────────────────────────
    target_price = round(entry_price + ATR_TARGET_MULT * atr, 2)
    stop_loss    = round(entry_price - ATR_STOP_MULT  * atr, 2)

    expected_profit_pct = round((target_price - entry_price) / entry_price * 100.0, 2)
    risk_pct            = round((entry_price - stop_loss) / entry_price * 100.0, 2)
    rr_ratio            = round(expected_profit_pct / risk_pct, 2) if risk_pct > 0 else None

    # ── Holding days estimate (multi-factor model) ────────────────────────
    # Factor 1: Volatility (ATR as % of price) → base duration
    atr_pct = (atr / entry_price) * 100.0 if entry_price > 0 else 2.0
    if atr_pct >= 4.0:
        base_days = 8       # high volatility → moves fast
    elif atr_pct >= 2.5:
        base_days = 12
    elif atr_pct >= 1.5:
        base_days = 16
    else:
        base_days = 20      # low volatility → needs more time

    # Factor 2: Trend strength (ADX) → speed multiplier
    if adx is not None and adx >= 35:
        adx_mult = 0.65     # strong trend → target reached faster
    elif adx is not None and adx >= 25:
        adx_mult = 0.80
    elif adx is not None and adx >= 15:
        adx_mult = 1.0
    else:
        adx_mult = 1.25     # no trend → slower

    # Factor 3: Conviction (CWAF score) → hold-longer multiplier
    if orchestrator_score >= 78:
        conv_mult = 1.15
    elif orchestrator_score >= 70:
        conv_mult = 1.08
    else:
        conv_mult = 1.0

    indicator_days = round(base_days * adx_mult * conv_mult)

    # Factor 4: Pattern-characteristic durations (bullish patterns only)
    pattern_day_vals: List[float] = []
    for pat in (patterns or []):
        if pat.get("direction") == "bullish":
            char = _PATTERN_CHAR_DURATIONS.get(pat.get("name", ""))
            if char:
                conf = pat.get("confidence", 0.5)
                pattern_day_vals.append(((char[0] + char[1]) / 2.0) * conf)

    if pattern_day_vals:
        avg_pat_days = round(sum(pattern_day_vals) / len(pattern_day_vals))
        # Blend: 60 % indicator-based, 40 % pattern-based
        holding_days_est = round(0.6 * indicator_days + 0.4 * avg_pat_days)
    else:
        holding_days_est = indicator_days

    holding_days_est = max(3, min(25, holding_days_est))

    # ── Trade-duration filter ─────────────────────────────────────────────
    if holding_days_est <= 2:
        return None

    # ── Exit date ─────────────────────────────────────────────────────────
    exit_date_est = signal_date + timedelta(days=int(holding_days_est * 1.4))
    if exit_date_est > result_date:
        exit_date_est = result_date

    # ── Profit probability  (CWAF confidence mapped → 40–85 %) ───────────
    profit_prob = round(min(85.0, max(40.0, final_confidence * 80.0 + 20.0)), 1)

    return {
        "direction":            "LONG",
        "entry_date":           signal_date.isoformat(),
        "exit_date_est":        exit_date_est.isoformat(),
        "result_date":          result_date.isoformat(),
        "entry_price":          round(entry_price, 2),
        "target_price":         target_price,
        "stop_loss":            stop_loss,
        "expected_profit_pct":  expected_profit_pct,
        "risk_pct":             risk_pct,
        "reward_risk_ratio":    rr_ratio,
        "confidence_score":     round(orchestrator_score, 1),
        "final_confidence":     round(final_confidence, 3),
        "profit_probability":   profit_prob,
        "trade_duration_days":  holding_days_est,
        "atr_at_entry":         round(atr, 2),
        "rsi_at_entry":         rsi,
        "adx_at_entry":         adx,
        # v3: transaction costs & slippage
        "transaction_cost_pct": 0.10,   # 10 bps round-trip (commission)
        "slippage_pct":         0.05,   # 5 bps estimated slippage per side
        "total_friction_pct":   0.20,   # 10 bps commission + 2×5 bps slippage
        "net_expected_profit_pct": round(expected_profit_pct - 0.20, 2),
        # v3: signal alignment from tech agent
        "signal_alignment":     _extract_signal_alignment(tech_result),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Pattern recognition extractor  (from raw tech result)
# ─────────────────────────────────────────────────────────────────────────────

def _extract_patterns(tech_result: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Pull chart pattern recognition details from the technical agent output."""
    if not tech_result:
        return []
    raw = tech_result.get("patterns", [])
    return [
        {
            "name":                 p.get("name"),
            "direction":            p.get("direction"),
            "confidence":           p.get("confidence"),
            "start_date":           p.get("start_date"),
            "end_date":             p.get("end_date"),
            "breakout_confirmed":   p.get("breakout_confirmed"),
            "volume_confirmation":  p.get("volume_confirmation"),
            "description":          p.get("description"),
        }
        for p in raw
    ]


def _extract_signal_alignment(tech_result: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Extract v3 signal alignment details from the technical agent output."""
    if not tech_result:
        return None
    sa = tech_result.get("signal_alignment")
    if not sa:
        return None
    return {
        "signal_count":       sa.get("signal_count", 0),
        "bullish_frameworks": sa.get("bullish_frameworks", 0),
        "entry_rules_met":    sa.get("entry_rules_met", 0),
        "confidence_pct":     sa.get("confidence_pct", 0),
        "confidence_label":   sa.get("confidence_label", "unknown"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Technical summary extractor
# ─────────────────────────────────────────────────────────────────────────────

def _extract_tech_summary(tech_result: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Extract key technical metrics from the raw tech result."""
    if not tech_result:
        return None

    exp = tech_result.get("experimental_score", {})
    ki  = tech_result.get("key_indicators", {}) or {}

    # Per-framework scores
    fw_scores: Dict[str, Any] = {}
    for name, fw in tech_result.get("frameworks", {}).items():
        fw_scores[name] = {
            "applicable": fw.get("applicable"),
            "score_pct":  fw.get("score_pct"),
        }

    return {
        "score":        exp.get("score") if isinstance(exp, dict) else None,
        "band":         exp.get("band") if isinstance(exp, dict) else None,
        "confidence":   exp.get("confidence") if isinstance(exp, dict) else None,
        "confidence_pct": exp.get("confidence_pct") if isinstance(exp, dict) else None,
        "subscores":    exp.get("subscores") if isinstance(exp, dict) else None,
        "signal_alignment": tech_result.get("signal_alignment"),
        "risk_management":  tech_result.get("risk_management"),
        "frameworks":   fw_scores,
        "key_indicators": {
            "close":    ki.get("close"),
            "ema_9":    ki.get("ema_9"),
            "ema_20":   ki.get("ema_20"),
            "ema_21":   ki.get("ema_21"),
            "ema_50":   ki.get("ema_50"),
            "ema_200":  ki.get("ema_200"),
            "rsi_14":   ki.get("rsi_14"),
            "macd":     ki.get("macd"),
            "macd_histogram": ki.get("macd_histogram"),
            "adx":      ki.get("adx"),
            "atr_14":   ki.get("atr_14"),
            "bb_pct_b": ki.get("bb_pct_b"),
            "obv_trend": ki.get("obv_trend"),
            "ichimoku_tenkan": ki.get("ichimoku_tenkan"),
            "ichimoku_kijun":  ki.get("ichimoku_kijun"),
            "roc_12":   ki.get("roc_12"),
            "cci_20":   ki.get("cci_20"),
            "cmf_20":   ki.get("cmf_20"),
            "supertrend_direction": ki.get("supertrend_direction"),
            "supertrend_line": ki.get("supertrend_line"),
            "stoch_rsi_k": ki.get("stoch_rsi_k"),
            "stoch_rsi_d": ki.get("stoch_rsi_d"),
            "kc_upper":  ki.get("kc_upper"),
            "kc_lower":  ki.get("kc_lower"),
            "squeeze_on": ki.get("squeeze_on"),
            "fibonacci": ki.get("fibonacci"),
            "market_structure": ki.get("market_structure"),
            "vwap_20":  ki.get("vwap_20"),
            "williams_r_14": ki.get("williams_r_14"),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Fundamental summary extractor
# ─────────────────────────────────────────────────────────────────────────────

def _extract_fund_summary(fund_result: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Extract key fundamental metrics from the raw fund result."""
    if not fund_result:
        return None

    exp = fund_result.get("experimental_score", {})
    dq  = fund_result.get("data_quality", {}) or {}

    # Per-framework scores
    fw_scores: Dict[str, Any] = {}
    for name, fw in fund_result.get("frameworks", {}).items():
        fw_scores[name] = {
            "applicable": fw.get("applicable"),
            "score_pct":  fw.get("score_pct"),
        }

    # Snapshot
    snap = fund_result.get("snapshot", {}) or {}

    return {
        "score":          exp.get("score") if isinstance(exp, dict) else None,
        "band":           exp.get("band") if isinstance(exp, dict) else None,
        "confidence":     exp.get("confidence") if isinstance(exp, dict) else None,
        "subscores":      exp.get("subscores") if isinstance(exp, dict) else None,
        "data_quality":   dq.get("coverage_quality"),
        "coverage_ratio": dq.get("coverage_ratio"),
        "frameworks":     fw_scores,
        "snapshot": {
            "market_cap_proxy":  snap.get("market_cap_proxy"),
            "revenue_current":   snap.get("revenue_current"),
            "eps_current":       snap.get("eps_current"),
            "eps_prior":         snap.get("eps_prior"),
            "net_income_current": snap.get("net_income_current"),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Actual optimal exit analysis  (peak price within the month)
# ─────────────────────────────────────────────────────────────────────────────

def _compute_actual_optimal_exit(
    ticker: str,
    entry_date: date,
    result_date: date,
    entry_price: float,
    target_price: float,
    stop_loss: float,
) -> Optional[Dict[str, Any]]:
    """
    Fetch daily bars for entry→result_date and compute:
      • optimal exit date (day of peak close)
      • optimal holding days
      • peak return %
      • target/stop hit dates
      • max drawdown from entry before exit
    """
    import pandas as pd

    # Polygon primary → yfinance fallback (same pattern as PolygonTechnicalClient)
    df: Optional[Any] = None
    if _bar_fetcher.is_available():
        df = _bar_fetcher.fetch_daily_bars(ticker, as_of=result_date, lookback_days=40)

    if df is None or df.empty:
        # yfinance fallback
        try:
            import yfinance as yf
            start = entry_date - timedelta(days=3)
            end   = result_date + timedelta(days=1)
            df = yf.download(
                ticker, start=start.isoformat(), end=end.isoformat(),
                auto_adjust=True, progress=False,
            )
            if df is not None and isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
        except Exception:
            df = None

    if df is None or df.empty:
        return None

    # Keep only bars from entry_date onward
    df = df[df.index >= pd.Timestamp(entry_date)]
    if df.empty:
        return None

    peak_price     = entry_price
    peak_date      = entry_date
    target_hit_dt: Optional[date] = None
    stop_hit_dt:   Optional[date] = None
    max_dd         = 0.0

    for idx_val, row in df.iterrows():
        bar_date = idx_val.date() if hasattr(idx_val, "date") else idx_val
        close = float(row["Close"])
        high  = float(row["High"])
        low   = float(row["Low"])

        if close > peak_price:
            peak_price = close
            peak_date  = bar_date

        if target_hit_dt is None and high >= target_price:
            target_hit_dt = bar_date

        if stop_hit_dt is None and low <= stop_loss:
            stop_hit_dt = bar_date

        dd = (close - entry_price) / entry_price * 100.0
        if dd < max_dd:
            max_dd = dd

    optimal_days = (peak_date - entry_date).days
    peak_return  = round((peak_price - entry_price) / entry_price * 100.0, 2)

    return {
        "optimal_exit_date":    peak_date.isoformat(),
        "optimal_holding_days": optimal_days,
        "peak_price":           round(peak_price, 2),
        "peak_return_pct":      peak_return,
        "target_hit":           target_hit_dt is not None,
        "target_hit_date":      target_hit_dt.isoformat() if target_hit_dt else None,
        "days_to_target":       (target_hit_dt - entry_date).days if target_hit_dt else None,
        "stop_hit":             stop_hit_dt is not None,
        "stop_hit_date":        stop_hit_dt.isoformat() if stop_hit_dt else None,
        "days_to_stop":         (stop_hit_dt - entry_date).days if stop_hit_dt else None,
        "max_drawdown_pct":     round(max_dd, 2),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Period runner  (single month for one ticker)
# ─────────────────────────────────────────────────────────────────────────────

def _run_period(ticker: str, signal_date: date, result_date: date) -> Dict[str, Any]:
    """Evaluate one (signal_date → result_date) window for *ticker*."""
    from agents.technical.service import analyze_ticker as tech_analyze
    from agents.fundamental.service import analyze_ticker as fund_analyze

    # ── Step 1: Run BOTH agents ───────────────────────────────────────────
    tech_result: Optional[Dict[str, Any]] = None
    tech_error:  Optional[str] = None
    fund_result: Optional[Dict[str, Any]] = None
    fund_error:  Optional[str] = None

    try:
        tech_result = tech_analyze(
            ticker=ticker,
            as_of_date=signal_date.isoformat(),
        )
    except Exception as exc:
        tech_error = str(exc)

    try:
        fund_result = fund_analyze(
            ticker=ticker,
            as_of_date=signal_date.isoformat(),
            data_source=_ORCH_SETTINGS.fund_data_source,
        )
    except Exception as exc:
        fund_error = str(exc)

    # ── Step 2: CWAF fusion ───────────────────────────────────────────────
    fusion = fuse_signals(
        tech_result=tech_result,
        tech_error=tech_error,
        fund_result=fund_result,
        fund_error=fund_error,
        settings=_ORCH_SETTINGS,
    )

    # ── Step 3: Compute LONG trade setup ──────────────────────────────────
    trade_setup = _compute_long_trade_setup(
        tech_result=tech_result,
        fund_result=fund_result,
        fusion_signal=fusion.final_signal,
        orchestrator_score=fusion.orchestrator_score,
        final_confidence=fusion.final_confidence,
        signal_date=signal_date,
        result_date=result_date,
    )

    # ── Step 4: Extract pattern recognition ───────────────────────────────
    patterns = _extract_patterns(tech_result)

    # ── Step 5: Extract agent summaries ───────────────────────────────────
    tech_summary = _extract_tech_summary(tech_result)
    fund_summary = _extract_fund_summary(fund_result)

    # ── Step 6: Actual period outcome ─────────────────────────────────────
    actual_outcome: Optional[Dict[str, Any]] = None
    start_price = None
    for src in (tech_result, fund_result):
        if src and "as_of_price" in src:
            start_price = src["as_of_price"]["price"]
            break

    if start_price:
        try:
            end_bar    = _price_client.get_price_as_of(ticker, result_date)
            end_price  = end_bar.close
            return_pct = round((end_price - start_price) / start_price * 100.0, 2)
            actual_dir = "UP" if return_pct >= 0 else "DOWN"

            # For LONG trades: correct if price went UP
            signal_correct: Optional[bool] = None
            if trade_setup:
                signal_correct = actual_dir == "UP"

            actual_outcome = {
                "end_price":        round(end_price, 2),
                "end_price_date":   end_bar.bar_date.isoformat(),
                "price_return_pct": return_pct,
                "actual_direction": actual_dir,
                "signal_correct":   signal_correct,
            }
        except Exception as exc:
            actual_outcome = {"error": str(exc)}

    # Optimal exit analysis (LONG trades only) — find peak within the month
    optimal_exit: Optional[Dict[str, Any]] = None
    if trade_setup and start_price:
        try:
            optimal_exit = _compute_actual_optimal_exit(
                ticker       = ticker,
                entry_date   = signal_date,
                result_date  = result_date,
                entry_price  = start_price,
                target_price = trade_setup["target_price"],
                stop_loss    = trade_setup["stop_loss"],
            )
        except Exception:
            optimal_exit = None

    return {
        "month":              signal_date.strftime("%B %Y"),
        "signal_date":        signal_date.isoformat(),
        "result_date":        result_date.isoformat(),
        # Orchestrator (CWAF) output
        "orchestrator_signal": fusion.final_signal,
        "orchestrator_score":  round(fusion.orchestrator_score, 1),
        "final_confidence":    round(fusion.final_confidence, 3),
        "conflict_detected":   fusion.conflict_detected,
        "conflict_resolution": fusion.conflict_resolution,
        "weights_applied":     fusion.weights_applied,
        # Trade setup (None if not LONG or filtered out)
        "trade_setup":         trade_setup,
        # Pattern recognition
        "patterns_detected":   patterns,
        "patterns_count":      len(patterns),
        # Agent summaries (for ML feature engineering)
        "tech_summary":        tech_summary,
        "fund_summary":        fund_summary,
        "tech_error":          tech_error,
        "fund_error":          fund_error,
        # Actual outcome
        "actual_outcome":      actual_outcome,
        # Optimal exit analysis (peak within the month for LONG trades)
        "optimal_exit_analysis": optimal_exit,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Post-backtest learning  (per-ticker profile from 12 months)
# ─────────────────────────────────────────────────────────────────────────────

def _learn_ticker_profile(
    periods: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Analyse all 12 months for one ticker to learn:
      - actual optimal holding duration vs. predicted
      - target/stop hit rates
      - seasonal pattern (which months were winning)
      - per-pattern performance
    """
    optimal_days_list:   List[int]   = []
    predicted_days_list: List[int]   = []
    actual_returns:      List[float] = []
    peak_returns:        List[float] = []
    max_dds:             List[float] = []
    target_hits = 0
    stop_hits   = 0
    winning_months: List[str] = []
    losing_months:  List[str] = []
    pattern_perf: Dict[str, Dict] = {}
    total_months = len(periods)

    for p in periods:
        ts  = p.get("trade_setup")
        ao  = p.get("actual_outcome") or {}
        oea = p.get("optimal_exit_analysis") or {}

        if not ts:
            continue

        if ts.get("trade_duration_days") is not None:
            predicted_days_list.append(ts["trade_duration_days"])

        if oea.get("optimal_holding_days") is not None:
            optimal_days_list.append(oea["optimal_holding_days"])
        if oea.get("peak_return_pct") is not None:
            peak_returns.append(oea["peak_return_pct"])
        if oea.get("max_drawdown_pct") is not None:
            max_dds.append(oea["max_drawdown_pct"])
        if oea.get("target_hit"):
            target_hits += 1
        if oea.get("stop_hit"):
            stop_hits += 1

        if ao.get("price_return_pct") is not None:
            actual_returns.append(ao["price_return_pct"])

        if ao.get("signal_correct") is True:
            winning_months.append(p["signal_date"][:7])
        elif ao.get("signal_correct") is False:
            losing_months.append(p["signal_date"][:7])

        for pat in p.get("patterns_detected", []):
            name = pat.get("name")
            if not name:
                continue
            if name not in pattern_perf:
                pattern_perf[name] = {"count": 0, "bullish": 0, "returns": []}
            pattern_perf[name]["count"] += 1
            if pat.get("direction") == "bullish":
                pattern_perf[name]["bullish"] += 1
            if ao.get("price_return_pct") is not None:
                pattern_perf[name]["returns"].append(ao["price_return_pct"])

    n_trades = len(predicted_days_list)

    def _safe_avg(lst):
        return round(sum(lst) / len(lst), 1) if lst else None

    avg_optimal  = _safe_avg(optimal_days_list)
    med_optimal  = (
        round(sorted(optimal_days_list)[len(optimal_days_list) // 2], 1)
        if optimal_days_list else None
    )
    avg_predict  = _safe_avg(predicted_days_list)
    dur_error    = round(abs(avg_predict - avg_optimal), 1) if avg_optimal is not None and avg_predict is not None else None
    avg_return   = _safe_avg(actual_returns)
    avg_peak     = _safe_avg(peak_returns)
    avg_dd       = _safe_avg(max_dds)

    # Pattern summary
    pat_summary: Dict[str, Dict] = {}
    best_pat_name:   Optional[str]   = None
    best_pat_return: float           = -999.0
    for name, stats in pattern_perf.items():
        r = stats["returns"]
        avg_r = round(sum(r) / len(r), 2) if r else None
        pat_summary[name] = {
            "appearances":              stats["count"],
            "bullish_count":            stats["bullish"],
            "avg_return_when_present":  avg_r,
        }
        if avg_r is not None and avg_r > best_pat_return:
            best_pat_return = avg_r
            best_pat_name   = name

    learned_duration = avg_optimal if avg_optimal is not None else (avg_predict if avg_predict is not None else 10)

    return {
        "total_months":                   total_months,
        "total_long_trades":              n_trades,
        "win_rate_pct":                   round(len(winning_months) / n_trades * 100, 1) if n_trades else None,
        "avg_optimal_holding_days":       avg_optimal,
        "median_optimal_holding_days":    med_optimal,
        "avg_predicted_holding_days":     avg_predict,
        "duration_prediction_error_days": dur_error,
        "target_hit_rate_pct":            round(target_hits / n_trades * 100, 1) if n_trades else None,
        "stop_hit_rate_pct":              round(stop_hits / n_trades * 100, 1) if n_trades else None,
        "avg_actual_return_pct":          avg_return,
        "avg_peak_return_pct":            avg_peak,
        "avg_max_drawdown_pct":           avg_dd,
        "winning_months":                 winning_months,
        "losing_months":                  losing_months,
        "pattern_performance":            pat_summary,
        "best_pattern":                   best_pat_name,
        "best_pattern_avg_return_pct":    best_pat_return if best_pat_name else None,
        "learned_optimal_duration":       round(learned_duration, 1),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Forward prediction  (Q1 + Q2 2026 based on learned 2025 profile)
# ─────────────────────────────────────────────────────────────────────────────

def _predict_forward(
    ticker: str,
    sector: str,
    cap_tier: str,
    learned: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Predict entry/exit dates for future months using the learned profile.
    Inputs come from :func:`_learn_ticker_profile`.  Predictions cover
    PREDICTION_MONTHS (Jan-Jun 2026).
    """
    n_trades      = learned.get("total_long_trades", 0)
    n_months      = learned.get("total_months", 12)
    bullish_rate  = round(n_trades / n_months * 100, 1) if n_months else 0
    learned_dur   = learned.get("learned_optimal_duration", 10)
    avg_return    = learned.get("avg_actual_return_pct")
    avg_peak      = learned.get("avg_peak_return_pct")
    target_rate   = learned.get("target_hit_rate_pct")
    avg_dd        = learned.get("avg_max_drawdown_pct")
    win_rate      = learned.get("win_rate_pct")

    # Seasonal: which month-numbers had winning trades in 2025?
    win_month_nums: set = set()
    for m in learned.get("winning_months", []):
        try:
            win_month_nums.add(int(m.split("-")[1]))
        except (ValueError, IndexError):
            pass

    predictions: List[Dict[str, Any]] = []
    for pred_start, pred_end in PREDICTION_MONTHS:
        month_num   = pred_start.month
        month_name  = pred_start.strftime("%B %Y")
        seasonal_ok = month_num in win_month_nums

        exit_date = pred_start + timedelta(days=int(learned_dur * 1.4))
        if exit_date > pred_end:
            exit_date = pred_end

        # Conviction based on consistency
        if n_trades >= 6 and (win_rate or 0) >= 60:
            conviction = "HIGH"
        elif n_trades >= 3 and (win_rate or 0) >= 40:
            conviction = "MODERATE"
        elif n_trades >= 1:
            conviction = "LOW"
        else:
            conviction = "NO DATA"

        predictions.append({
            "month":                       month_name,
            "predicted_entry_date":        pred_start.isoformat(),
            "predicted_exit_date":         exit_date.isoformat(),
            "predicted_holding_days":      round(learned_dur),
            "bullish_probability_pct":     bullish_rate,
            "expected_return_pct":         avg_return,
            "expected_peak_return_pct":    avg_peak,
            "target_hit_probability_pct":  target_rate,
            "max_historical_drawdown_pct": avg_dd,
            "seasonal_match":              seasonal_ok,
            "conviction":                  conviction,
            "historical_win_rate_pct":     win_rate,
        })

    return predictions


# ─────────────────────────────────────────────────────────────────────────────
# Ticker backtest runner  (all 12 months)
# ─────────────────────────────────────────────────────────────────────────────

def _run_ticker(
    ticker: str,
    sector: str,
    cap_tier: str,
    months: List[Tuple[date, date]],
) -> Dict[str, Any]:
    """Run all monthly periods for one ticker and build the result dict."""
    periods: List[Dict[str, Any]] = []
    for signal_date, result_date in months:
        period = _run_period(ticker, signal_date, result_date)
        periods.append(period)
        time.sleep(0.15)  # light rate-limit buffer

    # ── Summary stats ─────────────────────────────────────────────────────
    long_trades = [
        p for p in periods
        if p.get("actual_outcome")
        and isinstance((p.get("actual_outcome") or {}).get("signal_correct"), bool)
    ]
    correct  = sum(1 for p in long_trades if p["actual_outcome"]["signal_correct"])
    n_long   = len(long_trades)
    accuracy = round(correct / n_long * 100.0, 1) if n_long > 0 else None

    setups = [p["trade_setup"] for p in periods if p.get("trade_setup")]
    def _avg(key: str) -> Optional[float]:
        vals = [s[key] for s in setups if s.get(key) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    # Pattern stats
    all_patterns = []
    for p in periods:
        for pat in p.get("patterns_detected", []):
            all_patterns.append(pat)

    bullish_pats = sum(1 for p in all_patterns if p.get("direction") == "bullish")
    bearish_pats = sum(1 for p in all_patterns if p.get("direction") == "bearish")

    # Conflict frequency
    n_conflict = sum(1 for p in periods if p.get("conflict_detected"))

    # ── Learning & forward prediction ─────────────────────────────────────
    learned     = _learn_ticker_profile(periods)
    predictions = _predict_forward(ticker, sector, cap_tier, learned)

    return {
        "ticker":          ticker.upper(),
        "sector":          sector,
        "market_cap_tier": cap_tier,
        "periods":         periods,
        "summary": {
            "total_months":             len(periods),
            "long_signals":             n_long,
            "correct_signals":          correct,
            "accuracy_pct":             accuracy,
            "avg_expected_profit_pct":  _avg("expected_profit_pct"),
            "avg_confidence_score":     _avg("confidence_score"),
            "avg_final_confidence":     _avg("final_confidence"),
            "avg_profit_probability":   _avg("profit_probability"),
            "avg_trade_duration_days":  _avg("trade_duration_days"),
            "total_patterns_detected":  len(all_patterns),
            "bullish_patterns":         bullish_pats,
            "bearish_patterns":         bearish_pats,
            "conflict_months":          n_conflict,
        },
        "learned_profile":      learned,
        "forward_predictions":  predictions,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

# -- Trade sheet headers --
TRADE_HEADERS = [
    "Ticker", "Sector", "Cap Tier", "Month",
    "Entry Date", "Exit Date Est",
    "Direction", "Entry Price", "Target Price", "Stop Loss",
    "Expected Profit %", "Risk %", "R:R Ratio",
    "CWAF Score", "CWAF Confidence", "Profit Probability %",
    "Trade Duration (Days)",
    "ATR at Entry", "RSI at Entry", "ADX at Entry",
    # Agent scores
    "Tech Score", "Tech Band", "Fund Score", "Fund Band",
    "Conflict?", "Conflict Resolution",
    "Fund Data Quality",
    # Patterns inline
    "Patterns Count", "Pattern Names",
    # Actual
    "Actual End Price", "Actual Return %", "Actual Direction", "Signal Correct",
]

# -- Pattern log headers --
PATTERN_HEADERS = [
    "Ticker", "Sector", "Month",
    "Pattern Name", "Direction", "Confidence",
    "Start Date", "End Date",
    "Breakout Confirmed", "Volume Confirmation",
    "Description",
]

# -- Summary headers --
SUMMARY_HEADERS = [
    "Sector", "Ticker", "Cap Tier",
    "Long Signals", "Correct Signals", "Accuracy %",
    "Avg CWAF Score", "Avg CWAF Confidence", "Avg Profit Prob %",
    "Avg Expected Profit %", "Avg Trade Duration (Days)",
    "Patterns (Bull/Bear)", "Conflict Months",
]

# -- ML feature headers (wide format, one row per month per ticker) --
ML_HEADERS = [
    "Ticker", "Sector", "Cap Tier", "Month", "Signal Date", "Result Date",
    # Orchestrator
    "Orch Signal", "Orch Score", "CWAF Confidence",
    "Conflict", "Weights Tech", "Weights Fund",
    # Tech
    "Tech Score", "Tech Band",
    "EMA Trend", "MACD System", "RSI Regime", "Bollinger",
    "Volume OBV", "ADX Stochastic", "Pattern Recognition", "Ichimoku", "Momentum",
    "Close", "EMA 20", "EMA 50", "EMA 200",
    "RSI 14", "MACD Val", "MACD Histogram", "ADX", "ATR 14",
    "BB %B", "OBV Trend", "ROC 12", "CCI 20", "CMF 20",
    # Fund
    "Fund Score", "Fund Band", "Fund Data Quality",
    "Financial Health", "Valuation", "Quality", "Growth",
    "Revenue Current", "EPS Current", "EPS Prior",
    # Patterns
    "Patterns Count", "Bullish Patterns", "Bearish Patterns",
    # Optimal exit analysis
    "Optimal Holding Days", "Peak Return %", "Target Hit", "Days to Target",
    # Target
    "Actual Return %", "Actual Direction",
]

# -- Learned profiles headers --
LEARNED_HEADERS = [
    "Ticker", "Sector", "Cap Tier",
    "Long Trades", "Win Rate %",
    "Avg Optimal Days", "Median Optimal Days",
    "Avg Predicted Days", "Duration Error (Days)",
    "Target Hit Rate %", "Stop Hit Rate %",
    "Avg Actual Return %", "Avg Peak Return %",
    "Avg Max Drawdown %",
    "Best Pattern", "Best Pattern Return %",
    "Learned Duration (Days)",
]

# -- Forward predictions headers --
PREDICTION_HEADERS = [
    "Ticker", "Sector", "Cap Tier",
    "Month", "Predicted Entry", "Predicted Exit",
    "Holding Days",
    "Bullish Probability %",
    "Expected Return %", "Expected Peak Return %",
    "Target Hit Probability %",
    "Historical Max Drawdown %",
    "Seasonal Match", "Conviction", "Win Rate %",
]


def _build_trade_row(td: Dict, period: Dict) -> Optional[Dict]:
    """Build a flat trade row from a period dict, or None if no trade setup."""
    ts = period.get("trade_setup")
    if ts is None:
        return None

    ao  = period.get("actual_outcome") or {}
    tsu = period.get("tech_summary") or {}
    fsu = period.get("fund_summary") or {}
    pats = period.get("patterns_detected", [])

    pat_names = ", ".join(p.get("name", "") for p in pats) if pats else ""

    return {
        "Ticker":              td["ticker"],
        "Sector":              td["sector"],
        "Cap Tier":            td["market_cap_tier"],
        "Month":               period["month"],
        "Entry Date":          ts.get("entry_date"),
        "Exit Date Est":       ts.get("exit_date_est"),
        "Direction":           ts.get("direction"),
        "Entry Price":         ts.get("entry_price"),
        "Target Price":        ts.get("target_price"),
        "Stop Loss":           ts.get("stop_loss"),
        "Expected Profit %":   ts.get("expected_profit_pct"),
        "Risk %":              ts.get("risk_pct"),
        "R:R Ratio":           ts.get("reward_risk_ratio"),
        "CWAF Score":          ts.get("confidence_score"),
        "CWAF Confidence":     ts.get("final_confidence"),
        "Profit Probability %": ts.get("profit_probability"),
        "Trade Duration (Days)": ts.get("trade_duration_days"),
        "ATR at Entry":        ts.get("atr_at_entry"),
        "RSI at Entry":        ts.get("rsi_at_entry"),
        "ADX at Entry":        ts.get("adx_at_entry"),
        "Tech Score":          tsu.get("score"),
        "Tech Band":           tsu.get("band"),
        "Fund Score":          fsu.get("score"),
        "Fund Band":           fsu.get("band"),
        "Conflict?":           period.get("conflict_detected"),
        "Conflict Resolution": period.get("conflict_resolution"),
        "Fund Data Quality":   fsu.get("data_quality"),
        "Patterns Count":      period.get("patterns_count", 0),
        "Pattern Names":       pat_names,
        "Actual End Price":    ao.get("end_price"),
        "Actual Return %":     ao.get("price_return_pct"),
        "Actual Direction":    ao.get("actual_direction"),
        "Signal Correct":      ao.get("signal_correct"),
    }


def _build_pattern_row(td: Dict, period: Dict, pat: Dict) -> Dict:
    """Build a single pattern log row."""
    return {
        "Ticker":               td["ticker"],
        "Sector":               td["sector"],
        "Month":                period["month"],
        "Pattern Name":         pat.get("name"),
        "Direction":            pat.get("direction"),
        "Confidence":           pat.get("confidence"),
        "Start Date":           pat.get("start_date"),
        "End Date":             pat.get("end_date"),
        "Breakout Confirmed":   pat.get("breakout_confirmed"),
        "Volume Confirmation":  pat.get("volume_confirmation"),
        "Description":          pat.get("description"),
    }


def _build_ml_row(td: Dict, period: Dict) -> Dict:
    """Build one ML-ready feature row (every month, every ticker)."""
    tsu  = period.get("tech_summary") or {}
    fsu  = period.get("fund_summary") or {}
    ao   = period.get("actual_outcome") or {}
    oea  = period.get("optimal_exit_analysis") or {}
    pats = period.get("patterns_detected", [])
    t_fw = tsu.get("frameworks", {})
    t_ki = tsu.get("key_indicators", {})
    t_ss = tsu.get("subscores") or {}
    f_ss = fsu.get("subscores") or {}
    f_sn = fsu.get("snapshot") or {}
    wt   = period.get("weights_applied") or {}

    bullish_pat_count = sum(1 for p in pats if p.get("direction") == "bullish")
    bearish_pat_count = sum(1 for p in pats if p.get("direction") == "bearish")

    def _fw_score(name: str) -> Optional[float]:
        fw = t_fw.get(name, {})
        return fw.get("score_pct") if fw.get("applicable") else None

    return {
        "Ticker":          td["ticker"],
        "Sector":          td["sector"],
        "Cap Tier":        td["market_cap_tier"],
        "Month":           period["month"],
        "Signal Date":     period["signal_date"],
        "Result Date":     period["result_date"],
        "Orch Signal":     period.get("orchestrator_signal"),
        "Orch Score":      period.get("orchestrator_score"),
        "CWAF Confidence": period.get("final_confidence"),
        "Conflict":        period.get("conflict_detected"),
        "Weights Tech":    wt.get("tech"),
        "Weights Fund":    wt.get("fund"),
        "Tech Score":      tsu.get("score"),
        "Tech Band":       tsu.get("band"),
        "EMA Trend":       _fw_score("ema_trend"),
        "MACD System":     _fw_score("macd_system"),
        "RSI Regime":      _fw_score("rsi_regime"),
        "Bollinger":       _fw_score("bollinger"),
        "Volume OBV":      _fw_score("volume_obv"),
        "ADX Stochastic":  _fw_score("adx_stochastic"),
        "Pattern Recognition": _fw_score("pattern_recognition"),
        "Ichimoku":        _fw_score("ichimoku"),
        "Momentum":        _fw_score("momentum"),
        # v3 framework scores
        "Supertrend":      _fw_score("supertrend"),
        "Volatility Squeeze": _fw_score("volatility_squeeze"),
        "Entry/Exit Rules": _fw_score("entry_exit_rules"),
        "Close":           t_ki.get("close"),
        "EMA 9":           t_ki.get("ema_9"),
        "EMA 20":          t_ki.get("ema_20"),
        "EMA 21":          t_ki.get("ema_21"),
        "EMA 50":          t_ki.get("ema_50"),
        "EMA 200":         t_ki.get("ema_200"),
        "RSI 14":          t_ki.get("rsi_14"),
        "MACD Val":        t_ki.get("macd"),
        "MACD Histogram":  t_ki.get("macd_histogram"),
        "ADX":             t_ki.get("adx"),
        "ATR 14":          t_ki.get("atr_14"),
        "BB %B":           t_ki.get("bb_pct_b"),
        "OBV Trend":       t_ki.get("obv_trend"),
        "ROC 12":          t_ki.get("roc_12"),
        "CCI 20":          t_ki.get("cci_20"),
        "CMF 20":          t_ki.get("cmf_20"),
        # v3 indicator values
        "Supertrend Dir":  t_ki.get("supertrend_direction"),
        "Stoch RSI K":     t_ki.get("stoch_rsi_k"),
        "Stoch RSI D":     t_ki.get("stoch_rsi_d"),
        "Squeeze ON":      t_ki.get("squeeze_on"),
        "Market Structure": t_ki.get("market_structure"),
        "Williams %R":     t_ki.get("williams_r_14"),
        "VWAP 20":         t_ki.get("vwap_20"),
        # v3 signal alignment
        "Signal Count":    (tsu.get("signal_alignment") or {}).get("signal_count"),
        "Bullish FW Count": (tsu.get("signal_alignment") or {}).get("bullish_frameworks"),
        "Entry Rules Met": (tsu.get("signal_alignment") or {}).get("entry_rules_met"),
        "Confidence Pct":  (tsu.get("signal_alignment") or {}).get("confidence_pct"),
        "Confidence Label": (tsu.get("signal_alignment") or {}).get("confidence_label"),
        # v3 risk management
        "Stop Loss":       (tsu.get("risk_management") or {}).get("stop_loss"),
        "Take Profit 1":   (tsu.get("risk_management") or {}).get("take_profit_1"),
        "Take Profit 2":   (tsu.get("risk_management") or {}).get("take_profit_2"),
        "ATR %":           (tsu.get("risk_management") or {}).get("atr_pct"),
        # v3 transaction costs
        "Transaction Cost %": 0.10,
        "Slippage %":         0.10,
        "Total Friction %":   0.20,
        "Fund Score":      fsu.get("score"),
        "Fund Band":       fsu.get("band"),
        "Fund Data Quality": fsu.get("data_quality"),
        "Financial Health": f_ss.get("financial_health"),
        "Valuation":       f_ss.get("valuation"),
        "Quality":         f_ss.get("quality"),
        "Growth":          f_ss.get("growth"),
        "Revenue Current": f_sn.get("revenue_current"),
        "EPS Current":     f_sn.get("eps_current"),
        "EPS Prior":       f_sn.get("eps_prior"),
        "Patterns Count":  len(pats),
        "Bullish Patterns": bullish_pat_count,
        "Bearish Patterns": bearish_pat_count,
        "Optimal Holding Days": oea.get("optimal_holding_days"),
        "Peak Return %":        oea.get("peak_return_pct"),
        "Target Hit":           oea.get("target_hit"),
        "Days to Target":       oea.get("days_to_target"),
        "Actual Return %":      ao.get("price_return_pct"),
        "Actual Direction":     ao.get("actual_direction"),
    }


def _build_learned_row(td: Dict, lp: Dict) -> Dict:
    """Build one row for the Learned Profiles sheet."""
    return {
        "Ticker":                td["ticker"],
        "Sector":                td["sector"],
        "Cap Tier":              td["market_cap_tier"],
        "Long Trades":           lp.get("total_long_trades"),
        "Win Rate %":            lp.get("win_rate_pct"),
        "Avg Optimal Days":      lp.get("avg_optimal_holding_days"),
        "Median Optimal Days":   lp.get("median_optimal_holding_days"),
        "Avg Predicted Days":    lp.get("avg_predicted_holding_days"),
        "Duration Error (Days)": lp.get("duration_prediction_error_days"),
        "Target Hit Rate %":     lp.get("target_hit_rate_pct"),
        "Stop Hit Rate %":       lp.get("stop_hit_rate_pct"),
        "Avg Actual Return %":   lp.get("avg_actual_return_pct"),
        "Avg Peak Return %":     lp.get("avg_peak_return_pct"),
        "Avg Max Drawdown %":    lp.get("avg_max_drawdown_pct"),
        "Best Pattern":          lp.get("best_pattern"),
        "Best Pattern Return %": lp.get("best_pattern_avg_return_pct"),
        "Learned Duration (Days)": lp.get("learned_optimal_duration"),
    }


def _build_prediction_row(td: Dict, fp: Dict) -> Dict:
    """Build one row for the Forward Predictions sheet."""
    return {
        "Ticker":                    td["ticker"],
        "Sector":                    td["sector"],
        "Cap Tier":                  td["market_cap_tier"],
        "Month":                     fp.get("month"),
        "Predicted Entry":           fp.get("predicted_entry_date"),
        "Predicted Exit":            fp.get("predicted_exit_date"),
        "Holding Days":              fp.get("predicted_holding_days"),
        "Bullish Probability %":     fp.get("bullish_probability_pct"),
        "Expected Return %":         fp.get("expected_return_pct"),
        "Expected Peak Return %":    fp.get("expected_peak_return_pct"),
        "Target Hit Probability %":  fp.get("target_hit_probability_pct"),
        "Historical Max Drawdown %": fp.get("max_historical_drawdown_pct"),
        "Seasonal Match":            fp.get("seasonal_match"),
        "Conviction":                fp.get("conviction"),
        "Win Rate %":                fp.get("historical_win_rate_pct"),
    }


def _export_excel(all_results: List[Dict], output_path: Path) -> None:
    """Write a multi-sheet Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠  openpyxl not installed — skipping Excel export.")
        print("     Run:  pip install openpyxl")
        return

    # Colour palettes
    FILL_HEADER_GREEN  = PatternFill("solid", fgColor="1E8449")
    FILL_HEADER_BLUE   = PatternFill("solid", fgColor="154360")
    FILL_HEADER_ORANGE = PatternFill("solid", fgColor="784212")
    FILL_HEADER_PURPLE = PatternFill("solid", fgColor="6C3483")
    FILL_HEADER_TEAL   = PatternFill("solid", fgColor="148F77")
    FONT_HEADER        = Font(bold=True, color="FFFFFF", size=10)
    FILL_CORRECT       = PatternFill("solid", fgColor="A9DFBF")
    FILL_WRONG         = PatternFill("solid", fgColor="F5B7B1")
    FILL_LONG          = PatternFill("solid", fgColor="EAF2FF")
    ALIGN_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ── Collect rows ──────────────────────────────────────────────────────
    trade_rows:    List[Dict] = []
    pattern_rows:  List[Dict] = []
    ml_rows:       List[Dict] = []
    sector_trades: Dict[str, List[Dict]] = {}

    for td in all_results:
        for p in td["periods"]:
            # Trade row (LONG only)
            row = _build_trade_row(td, p)
            if row:
                trade_rows.append(row)
                sector_trades.setdefault(td["sector"], []).append(row)

            # Pattern rows (all months)
            for pat in p.get("patterns_detected", []):
                pattern_rows.append(_build_pattern_row(td, p, pat))

            # ML rows (every month)
            ml_rows.append(_build_ml_row(td, p))

    def _write_sheet(ws, headers: List[str], rows: List[Dict], fill: PatternFill,
                     colour_correct: bool = False) -> None:
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font  = FONT_HEADER
            cell.fill  = fill
            cell.alignment = ALIGN_CENTER

        for ri, row in enumerate(rows, 2):
            correct = row.get("Signal Correct")
            for ci, header in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(header))
                cell.alignment = ALIGN_CENTER
                if colour_correct:
                    if correct is True:
                        cell.fill = FILL_CORRECT
                    elif correct is False:
                        cell.fill = FILL_WRONG
                    else:
                        cell.fill = FILL_LONG

        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16
        ws.freeze_panes = "A2"

    # ── Workbook ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # -- Summary sheet -------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    for ci, header in enumerate(SUMMARY_HEADERS, 1):
        cell = ws_sum.cell(row=1, column=ci, value=header)
        cell.font   = FONT_HEADER
        cell.fill   = FILL_HEADER_ORANGE
        cell.alignment = ALIGN_CENTER

    ri = 2
    for td in sorted(all_results, key=lambda x: (x["sector"], x["ticker"])):
        s = td["summary"]
        acc = s.get("accuracy_pct")
        pat_str = f"{s.get('bullish_patterns', 0)}/{s.get('bearish_patterns', 0)}"
        row_vals = [
            td["sector"], td["ticker"], td["market_cap_tier"],
            s.get("long_signals"), s.get("correct_signals"),
            f"{acc:.1f}%" if acc is not None else "N/A",
            s.get("avg_confidence_score"), s.get("avg_final_confidence"),
            s.get("avg_profit_probability"),
            s.get("avg_expected_profit_pct"), s.get("avg_trade_duration_days"),
            pat_str, s.get("conflict_months"),
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.alignment = ALIGN_CENTER
            if ci == 6 and acc is not None:
                cell.fill = FILL_CORRECT if acc >= 60 else (FILL_WRONG if acc < 45 else PatternFill())
        ri += 1

    for ci in range(1, len(SUMMARY_HEADERS) + 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 20
    ws_sum.freeze_panes = "A2"

    # -- All Trades sheet ----------------------------------------------------
    ws_all = wb.create_sheet("All Trades")
    _write_sheet(ws_all, TRADE_HEADERS, trade_rows, FILL_HEADER_GREEN, colour_correct=True)

    # -- Pattern Log sheet ---------------------------------------------------
    ws_pat = wb.create_sheet("Pattern Log")
    _write_sheet(ws_pat, PATTERN_HEADERS, pattern_rows, FILL_HEADER_PURPLE)

    # -- ML Features sheet ---------------------------------------------------
    ws_ml = wb.create_sheet("ML Features")
    _write_sheet(ws_ml, ML_HEADERS, ml_rows, FILL_HEADER_TEAL)

    # -- Learned Profiles sheet ----------------------------------------------
    learned_rows: List[Dict] = []
    for td in all_results:
        lp = td.get("learned_profile")
        if lp:
            learned_rows.append(_build_learned_row(td, lp))

    ws_lp = wb.create_sheet("Learned Profiles")
    _write_sheet(ws_lp, LEARNED_HEADERS, learned_rows, FILL_HEADER_ORANGE)

    # -- Forward Predictions sheet -------------------------------------------
    pred_rows: List[Dict] = []
    for td in all_results:
        for fp in td.get("forward_predictions", []):
            pred_rows.append(_build_prediction_row(td, fp))

    ws_fp = wb.create_sheet("Forward Predictions")
    _write_sheet(ws_fp, PREDICTION_HEADERS, pred_rows, FILL_HEADER_PURPLE)

    # -- Per-sector sheets ---------------------------------------------------
    for sector, rows in sorted(sector_trades.items()):
        sheet_name = sector[:31]
        ws_sec = wb.create_sheet(sheet_name)
        _write_sheet(ws_sec, TRADE_HEADERS, rows, FILL_HEADER_BLUE, colour_correct=True)

    wb.save(output_path)
    print(f"  ✓ Excel  → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI helpers
# ─────────────────────────────────────────────────────────────────────────────

def _print_header(n_sectors: int, n_tickers: int, n_months: int, workers: int) -> None:
    print()
    print("█" * 68)
    print("█  HALAL ORCHESTRATOR (CWAF) BACKTEST  —  Jan 2025 → Dec 2025")
    print("█  Technical + Fundamental Agents → CWAF Fusion → LONG only")
    print("█" * 68)
    print(f"  Sectors    : {n_sectors}")
    print(f"  Tickers    : {n_tickers}")
    print(f"  Months     : {n_months}  (Jan – Dec 2025)")
    print(f"  Data source: Polygon.io (OHLCV) + yfinance (fundamentals)")
    print(f"  Mode       : LONG trades only | trade_duration > 2 days")
    print(f"  Workers    : {workers}")
    print(f"  Output dir : {OUTPUT_DIR}")
    print()


def _print_sector_summary(all_results: List[Dict]) -> None:
    print()
    print("─" * 72)
    print(f"  {'SECTOR':<28}  {'TICKERS':>7}  {'LONGS':>7}  {'ACCURACY':>10}  {'PATTERNS':>8}")
    print("─" * 72)
    per_sector: Dict[str, Dict] = {}
    for td in all_results:
        sec = td["sector"]
        if sec not in per_sector:
            per_sector[sec] = {"n": 0, "dir": 0, "cor": 0, "pats": 0}
        per_sector[sec]["n"]    += 1
        per_sector[sec]["dir"]  += td["summary"]["long_signals"] or 0
        per_sector[sec]["cor"]  += td["summary"]["correct_signals"] or 0
        per_sector[sec]["pats"] += td["summary"]["total_patterns_detected"] or 0

    total_dir = total_cor = total_tickers = total_pats = 0
    for sec in sorted(per_sector):
        d = per_sector[sec]
        acc = f"{d['cor'] / d['dir'] * 100:.1f}%" if d["dir"] > 0 else "N/A"
        print(f"  {sec:<28}  {d['n']:>7}  {d['dir']:>7}  {acc:>10}  {d['pats']:>8}")
        total_dir     += d["dir"]
        total_cor     += d["cor"]
        total_tickers += d["n"]
        total_pats    += d["pats"]

    print("─" * 72)
    overall = f"{total_cor / total_dir * 100:.1f}%" if total_dir > 0 else "N/A"
    print(f"  {'OVERALL':<28}  {total_tickers:>7}  {total_dir:>7}  {overall:>10}  {total_pats:>8}")
    print("─" * 72)
    print()


# ─────────────────────────────────────────────────────────────────────────────
# Master JSON builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_master(
    all_results: List[Dict],
    months: List[Tuple[date, date]],
) -> Dict[str, Any]:
    """Aggregate all ticker results into a single master JSON."""

    all_trades: List[Dict]     = []
    all_patterns: List[Dict]   = []
    sector_agg: Dict[str, Dict] = {}

    for td in all_results:
        sec = td["sector"]
        if sec not in sector_agg:
            sector_agg[sec] = {
                "tickers":           [],
                "total_long":        0,
                "total_correct":     0,
                "total_patterns":    0,
            }
        sector_agg[sec]["tickers"].append(td["ticker"])
        sector_agg[sec]["total_long"]    += td["summary"]["long_signals"] or 0
        sector_agg[sec]["total_correct"] += td["summary"]["correct_signals"] or 0
        sector_agg[sec]["total_patterns"] += td["summary"]["total_patterns_detected"] or 0

        for p in td["periods"]:
            row = _build_trade_row(td, p)
            if row:
                all_trades.append(row)
            for pat in p.get("patterns_detected", []):
                all_patterns.append(_build_pattern_row(td, p, pat))

    for sec, agg in sector_agg.items():
        n = agg["total_long"]
        c = agg["total_correct"]
        agg["accuracy_pct"] = round(c / n * 100.0, 1) if n > 0 else None

    return {
        "metadata": {
            "backtest_type":         "CWAF Orchestrator (Technical + Fundamental)",
            "backtest_window":       "January 2025 – December 2025",
            "months":                [m[0].isoformat() for m in months],
            "total_tickers":         len(all_results),
            "sectors":               sorted(sector_agg.keys()),
            "total_long_trades":     len(all_trades),
            "total_patterns":        len(all_patterns),
            "data_source_ohlcv":     "Polygon.io (primary) / yfinance (fallback)",
            "data_source_fundamentals": "yfinance",
            "trade_mode":            "LONG only",
            "trade_duration_filter": "holding_days_est > 2",
            "atr_target_mult":       ATR_TARGET_MULT,
            "atr_stop_mult":         ATR_STOP_MULT,
            "generated_at":          date.today().isoformat(),
        },
        "sector_summary":    sector_agg,
        "all_trades":        all_trades,
        "all_patterns":      all_patterns,
        "ticker_summaries":   {td["ticker"]: td["summary"] for td in all_results},
        "learned_profiles":   {td["ticker"]: td.get("learned_profile") for td in all_results},
        "forward_predictions": {td["ticker"]: td.get("forward_predictions") for td in all_results},
    }


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Halal universe CWAF orchestrator backtest — 2025",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--sector",      default=None,   help="Run only this sector")
    parser.add_argument("--tickers",     default=None,   help="Comma-separated ticker overrides")
    parser.add_argument("--ticker-file", default=None,   help="JSON file with {sector: [{ticker, cap_tier, ...}]}")
    parser.add_argument("--per-sector",  type=int, default=10, help="Tickers per sector (default: 10)")
    parser.add_argument("--workers",     type=int, default=5,  help="Concurrent threads (default: 5)")
    parser.add_argument("--months",     type=int, default=12, help="Months to run 1-12 (default: 12)")
    parser.add_argument("--resume",     action="store_true",  help="Skip tickers already saved")
    parser.add_argument("--no-excel",   action="store_true",  help="Skip Excel export")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Choose months window ──────────────────────────────────────────────
    months = MONTHS_2025[: max(1, min(args.months, 12))]

    # ── Build ticker universe ─────────────────────────────────────────────
    if args.tickers:
        ticker_meta = [
            {
                "ticker":          t.strip().upper(),
                "name":            t.strip().upper(),
                "market_cap_tier": "large",
                "cap_display":     "large",
            }
            for t in args.tickers.split(",")
            if t.strip()
        ]
        universe: Dict[str, List[Dict]] = {"Custom": ticker_meta}
    elif args.ticker_file:
        # Load from curated JSON file (e.g. halal_top5_per_sector.json)
        tf = Path(args.ticker_file) if os.path.isabs(args.ticker_file) else ROOT / args.ticker_file
        with open(tf) as fh:
            raw = json.load(fh)
        universe = {}
        for sector, entries in raw.items():
            stocks = []
            for e in entries:
                ticker = e["ticker"] if isinstance(e, dict) else e
                cap    = e.get("cap_tier", "large") if isinstance(e, dict) else "large"
                stocks.append({
                    "ticker": ticker.upper(),
                    "name": ticker.upper(),
                    "market_cap_tier": cap,
                    "cap_display": cap,
                })
            if stocks:
                universe[sector] = stocks
    else:
        universe = _load_ticker_universe(per_sector=args.per_sector)

    if args.sector:
        universe = {
            k: v for k, v in universe.items()
            if k.lower() == args.sector.lower()
        }
        if not universe:
            print(f"  ERROR: sector '{args.sector}' not found in universe.")
            sys.exit(1)

    total_tickers = sum(len(v) for v in universe.values())
    _print_header(len(universe), total_tickers, len(months), args.workers)

    # ── Build task list ───────────────────────────────────────────────────
    tasks: List[Tuple[str, str, str]] = []
    all_results: List[Dict] = []

    if args.resume:
        for cache_path in sorted(OUTPUT_DIR.glob("*_halal_orch_2025.json")):
            if cache_path.stem.startswith("master"):
                continue
            try:
                with open(cache_path) as fh:
                    data = json.load(fh)
                    if "ticker" in data:
                        all_results.append(data)
            except (json.JSONDecodeError, KeyError):
                pass
        cached_tickers = {td["ticker"] for td in all_results}
        print(f"  Loaded {len(all_results)} cached results.", flush=True)
    else:
        cached_tickers: set = set()

    for sector, stocks in universe.items():
        for stock in stocks:
            ticker   = stock["ticker"].upper()
            cap_tier = stock.get("cap_display", stock.get("market_cap_tier", "unknown"))
            if args.resume and ticker in cached_tickers:
                print(f"    [{ticker:<8}] ← cache hit", flush=True)
                continue
            tasks.append((ticker, sector, cap_tier))

    if not tasks and not all_results:
        print("  Nothing to run. Exiting.")
        sys.exit(0)

    # ── Run backtest with thread pool ─────────────────────────────────────
    if tasks:
        print(f"\n  Running {len(tasks)} tickers × {len(months)} months ...\n")

    def _worker(task: Tuple[str, str, str]) -> Dict:
        ticker, sector, cap_tier = task
        print(f"  → {ticker:<8}  [{sector}] [{cap_tier}]", flush=True)
        t0   = time.time()
        data = _run_ticker(ticker, sector, cap_tier, months)
        elapsed = time.time() - t0
        s   = data["summary"]
        acc_str = f"{s['accuracy_pct']:.0f}%" if s.get("accuracy_pct") is not None else "N/A"
        pats = s.get("total_patterns_detected", 0)
        print(
            f"  ✓ {ticker:<8}  long={s['long_signals']}/{len(months)}"
            f"  acc={acc_str}  pats={pats}  ({elapsed:.0f}s)",
            flush=True,
        )
        return data

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(_worker, task): task for task in tasks}
        for fut in as_completed(futures):
            try:
                ticker_data = fut.result()
            except Exception as exc:
                task = futures[fut]
                print(f"  ✗ {task[0]:<8}  FAILED: {exc}", flush=True)
                continue

            all_results.append(ticker_data)

            # Save per-ticker JSON immediately
            cache_path = OUTPUT_DIR / f"{ticker_data['ticker']}_halal_orch_2025.json"
            with open(cache_path, "w") as fh:
                json.dump(ticker_data, fh, indent=2, default=str)

    # ── Write master JSON ─────────────────────────────────────────────────
    master = _build_master(all_results, months)
    master_path = OUTPUT_DIR / "master_halal_orch_2025.json"
    with open(master_path, "w") as fh:
        json.dump(master, fh, indent=2, default=str)
    print(f"\n  ✓ Master JSON → {master_path}")

    # ── Excel export ──────────────────────────────────────────────────────
    if not args.no_excel:
        xlsx_path = OUTPUT_DIR / "halal_orchestrator_backtest_2025.xlsx"
        print("  Building Excel workbook ...")
        _export_excel(all_results, xlsx_path)

    # ── Print summary table ───────────────────────────────────────────────
    _print_sector_summary(all_results)
    print(f"  Output → {OUTPUT_DIR}\n")


if __name__ == "__main__":
    main()
