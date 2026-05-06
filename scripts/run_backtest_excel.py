#!/usr/bin/env python3
"""
run_backtest_excel.py — Multi-threaded backtest runner with Excel output.

Selects up to --tickers-per-sector tickers from each sector in the halal
universe, runs the full TA + FA orchestrator pipeline for every ticker, and
writes results incrementally to a thread-safe Excel workbook.

Design rules
────────────
• Data is strictly bounded by --date (cutoff). No post-cutoff Polygon calls.
• Entry price is estimated from the last available bar (cutoff close).
• Entry conviction (IMMEDIATE / RETEST_ENTRY / WAIT_FOR_RETEST) is derived
  from price extension and staleness — never from future data.
• Each thread writes its result to Excel as soon as it finishes, using a
  file-level lock to prevent corruption.
• Re-running the script updates existing rows (matched by Ticker + Cutoff)
  rather than appending duplicates.

Usage
─────
  python scripts/run_backtest_excel.py
  python scripts/run_backtest_excel.py --date 2025-12-31 --workers 4
  python scripts/run_backtest_excel.py --date 2025-12-31 --workers 6 --target-days 20
  python scripts/run_backtest_excel.py --date 2025-12-31 --workers 4 --tickers-per-sector 10
  python scripts/run_backtest_excel.py --date 2025-12-31 --workers 4 --sector "Energy"

Parameters
──────────
  --date              Cutoff date (YYYY-MM-DD). Default: last weekday before today.
  --workers           Thread pool size. Default: 4. Recommended: 3–6.
  --target-days       Max trade window in trading days. Default: 20.
  --tickers-per-sector Max tickers sampled per sector. Default: 9 (~100 total).
  --sector            Run only this sector (optional filter).

Output
──────
  data/output/backtests/<YYYY-MM-DD>/
      <YYYY-MM-DD>.xlsx    — workbook, one sheet "Backtest Results"
      run_log.txt          — append-only plain-text log of this run
"""

from __future__ import annotations

import argparse
import json
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── project root on path ─────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

HALAL_JSON  = ROOT / "data" / "halal_universe" / "halal_sector_tickers.json"
OUTPUT_BASE = ROOT / "data" / "output" / "backtests"

# ── Excel columns (order matters) ────────────────────────────────────────────
HEADERS = [
    "Ticker", "Sector", "Cutoff Date", "Signal", "Confidence Score",
    "Tech Score", "Fund Score", "Pattern Name", "Pattern Start", "Pattern End",
    "Pattern Timeframe", "Breakout Date", "Breakout Price", "Days Since Breakout",
    "Entry Conviction", "Price Extension %",
    "Entry Date (Earliest)", "Entry Price (Est.)",
    "Retest Zone Low", "Retest Zone High",
    "Target Price", "Stop Loss", "R/R Ratio",
    "ATR", "ADX", "RSI",
    "Est. Days to Target", "Est. Target Date",
    "Outcome", "Exit Date", "Exit Price",
    "Gross P&L %", "Net P&L %",
    "No Trade Reason",
    "Run Timestamp",
]

# Outcome → fill colour (ARGB)
_OUTCOME_FILL: Dict[str, str] = {
    "HIT_TARGET":     "FF92D050",   # green
    "HIT_STOP":       "FFFF0000",   # red
    "EXPIRED":        "FF4472C4",   # blue
    "OPEN":           "FFFFC000",   # amber
    "SKIP":           "FFD9D9D9",   # grey  — criteria not met, no entry taken
}

# ── Thread-safe Excel writer ──────────────────────────────────────────────────
_excel_lock = threading.Lock()


def _ensure_workbook(path: Path) -> openpyxl.Workbook:
    """Load existing workbook or create a new one with formatted headers."""
    if path.exists():
        return openpyxl.load_workbook(path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Backtest Results"

    # Header row styling
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF203864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align

    # Column widths (approximate)
    col_widths = {
        1: 8,   # Ticker
        2: 22,  # Sector
        3: 13,  # Cutoff Date
        4: 11,  # Sentiment
        5: 12,  # Confidence
        6: 10,  # Tech Score
        7: 10,  # Fund Score
        8: 22,  # Pattern Name
        9: 13,  # Pattern Start
        10: 13, # Pattern End
        11: 12, # Pattern Timeframe
        12: 13, # Breakout Date
        13: 13, # Breakout Price
        14: 14, # Days Since Breakout
        15: 18, # Entry Conviction
        16: 14, # Price Extension %
        17: 16, # Entry Date
        18: 14, # Entry Price
        19: 14, # Retest Low
        20: 14, # Retest High
        21: 13, # Target
        22: 11, # Stop
        23: 9,  # R/R
        24: 8,  # ATR
        25: 8,  # ADX
        26: 8,  # RSI
        27: 14, # Est Days
        28: 15, # Est Target Date
        29: 14, # Outcome
        30: 13, # Exit Date
        31: 12, # Exit Price
        32: 12, # Gross P&L
        33: 11, # Net P&L
        34: 50, # No Trade Reason
        35: 18, # Run Timestamp
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    return wb


def _find_existing_row(ws, ticker: str, cutoff_date: str) -> Optional[int]:
    """Return 1-based row index if ticker+cutoff already exists, else None."""
    for row_idx in range(2, ws.max_row + 1):
        if (ws.cell(row_idx, 1).value == ticker and
                str(ws.cell(row_idx, 3).value) == cutoff_date):
            return row_idx
    return None


def _write_row(path: Path, row_values: List[Any], ticker: str, cutoff_date: str) -> None:
    """
    Thread-safe write: acquire lock → load workbook → upsert row → save.
    Ticker + Cutoff Date is the unique key. Existing rows are overwritten,
    new tickers are appended.
    """
    with _excel_lock:
        wb = _ensure_workbook(path)
        ws = wb.active

        existing_row = _find_existing_row(ws, ticker, cutoff_date)
        if existing_row:
            for col_idx, val in enumerate(row_values, start=1):
                ws.cell(row=existing_row, column=col_idx).value = val
            target_row = existing_row
        else:
            ws.append(row_values)
            target_row = ws.max_row

        # Colour the Outcome cell
        outcome_val = row_values[HEADERS.index("Outcome")]
        fill_hex    = _OUTCOME_FILL.get(str(outcome_val), "FFFFFFFF")
        outcome_col = HEADERS.index("Outcome") + 1
        ws.cell(target_row, outcome_col).fill = PatternFill("solid", fgColor=fill_hex)

        wb.save(path)


# ── Ticker selection ──────────────────────────────────────────────────────────

def _select_tickers(
    tickers_per_sector: int,
    sector_filter: Optional[str],
) -> List[Tuple[str, str]]:
    """
    Return list of (ticker, sector) tuples.
    Skips N/A sector and tickers with spaces or len > 6 (WI / malformed).
    Proportionally samples up to tickers_per_sector per sector.
    """
    raw: Dict[str, List[str]] = json.loads(HALAL_JSON.read_text())
    seen: set = set()
    result: List[Tuple[str, str]] = []

    for sector, tickers in raw.items():
        if sector == "N/A":
            continue
        if sector_filter and sector.lower() != sector_filter.lower():
            continue
        count = 0
        for t in tickers:
            if " " in t or len(t) > 6:
                continue
            if t in seen:
                continue
            seen.add(t)
            result.append((t, sector))
            count += 1
            if count >= tickers_per_sector:
                break

    return result


# ── Worker ────────────────────────────────────────────────────────────────────

def _predict_one(
    ticker: str,
    sector: str,
    cutoff_str: str,
    target_days: int,
    excel_path: Path,
    log_path: Path,
) -> str:
    """
    Run predict_trade() for one ticker, build the Excel row, write it.
    Returns a short status string for console progress.
    """
    from agents.technical.service import predict_trade

    ts = datetime.now().isoformat(timespec="seconds")

    try:
        result   = predict_trade(ticker=ticker, cutoff_date=cutoff_str, target_days=target_days)
        trade    = result.get("trade")
        sentiment = result.get("sentiment", "?")
        score     = result.get("confidence_score", 0)

        if trade:
            outcome   = trade.get("exit_outcome", "OPEN")
            conviction = trade.get("entry_conviction", "")
            row = [
                ticker,
                sector,
                cutoff_str,
                "BUY",     # Signal — entry criteria met, long trade taken
                round(score, 1),
                round(result.get("tech_score", 0), 1),
                round(result.get("fund_score", 0), 1),
                trade.get("pattern_name", ""),
                trade.get("pattern_start", ""),
                trade.get("pattern_end", ""),
                "Daily",   # timeframe — daily bars only currently
                trade.get("true_breakout_date", ""),
                trade.get("pattern_breakout_price", ""),
                trade.get("days_since_breakout", ""),
                conviction,
                trade.get("price_extension_pct", ""),
                trade.get("entry_earliest", trade.get("entry_date", "")),
                trade.get("entry_price", ""),
                trade.get("retest_zone_low", ""),
                trade.get("retest_zone_high", ""),
                trade.get("target_price", ""),
                trade.get("stop_loss", ""),
                trade.get("reward_risk_ratio", ""),
                trade.get("atr_at_entry", ""),
                trade.get("adx_at_entry", ""),
                trade.get("rsi_at_entry", ""),
                trade.get("estimated_days_to_target", ""),
                trade.get("estimated_target_date", ""),
                outcome,
                trade.get("exit_date", ""),
                trade.get("exit_price", ""),
                trade.get("gross_profit_pct", ""),
                trade.get("net_profit_pct", ""),
                "",    # no_trade_reason blank for trades
                ts,
            ]
            status = f"✓ {ticker:<6} [BUY    ] {conviction:<18} {outcome}"
        else:
            # ── SKIP row: entry criteria not met, but still compute projected
            # outcomes so the backtest has a complete feature row for ML.
            reason = result.get("no_trade_reason") or ""
            ki     = result.get("key_indicators") or {}
            pats   = result.get("patterns") or []

            skip_close = ki.get("close")
            skip_atr   = ki.get("atr_14")
            skip_adx   = ki.get("adx")
            skip_rsi   = ki.get("rsi_14")

            # Signal: AVOID for bearish, HOLD for neutral/mixed/bullish-but-no-entry
            signal = "AVOID" if sentiment == "bearish" else "HOLD"

            # Best pattern target for projected "if entered" target price
            skip_pat_tgt  = None
            skip_pat_name = ""
            if pats:
                first = pats[0]
                if isinstance(first, dict):
                    skip_pat_tgt  = first.get("pattern_target") or first.get("target")
                    skip_pat_name = first.get("name") or first.get("pattern_name") or ""

            if skip_close is not None and skip_atr is not None:
                skip_atr_f   = float(skip_atr)
                skip_close_f = float(skip_close)
                proj_stop    = round(skip_close_f - 2.0 * skip_atr_f, 2)
                proj_tgt     = round(float(skip_pat_tgt), 2) if skip_pat_tgt else round(skip_close_f + 3.0 * skip_atr_f, 2)
                denom        = skip_close_f - proj_stop
                proj_rr      = round((proj_tgt - skip_close_f) / denom, 2) if denom > 0 else ""
                proj_entry   = skip_close_f
                # % gain if price reaches target, % loss if it hits stop
                proj_gain_pct = round((proj_tgt - skip_close_f) / skip_close_f * 100, 2)
                proj_loss_pct = round((proj_stop - skip_close_f) / skip_close_f * 100, 2)
                # Estimated exit date: cutoff + target_days trading days ≈ ×1.4 calendar
                try:
                    from datetime import datetime as _dt
                    _cutoff_d  = _dt.strptime(cutoff_str, "%Y-%m-%d").date()
                    _exit_d    = _cutoff_d + timedelta(days=int(target_days * 1.4))
                    proj_exit_date = _exit_d.isoformat()
                except Exception:
                    proj_exit_date = ""
            else:
                proj_stop = proj_tgt = proj_rr = proj_entry = ""
                proj_gain_pct = proj_loss_pct = proj_exit_date = ""

            row = [
                ticker, sector, cutoff_str,
                signal, round(score, 1),              # Signal = HOLD / AVOID
                round(result.get("tech_score", 0), 1),
                round(result.get("fund_score", 0), 1),
                skip_pat_name, "", "", "",  # Pattern Name/Start/End/Timeframe
                "", "", "",                  # Breakout Date/Price/Days
                signal,                      # Entry Conviction = HOLD / AVOID
                "",                          # Price Extension %
                cutoff_str,                  # Entry Date (would-have-been)
                proj_entry,                  # Entry Price (projected)
                "", "",                      # Retest zone (N/A for SKIP)
                proj_tgt,                    # Target Price (projected)
                proj_stop,                   # Stop Loss (projected)
                proj_rr,                     # R/R (projected)
                skip_atr, skip_adx, skip_rsi,
                target_days,                 # Est Days to Target
                proj_exit_date,              # Est Target Date (projected)
                "SKIP",                      # Outcome
                proj_exit_date,              # Exit Date (projected)
                proj_tgt,                    # Exit Price (projected target)
                proj_gain_pct,               # Gross P&L % (to target)
                proj_loss_pct,               # Net P&L % (to stop — worst case)
                reason, ts,
            ]
            status = f"· {ticker:<6} [{signal:<6}] {reason[:50]}"

        _write_row(excel_path, row, ticker, cutoff_str)

        # Append to log
        with _excel_lock:
            with open(log_path, "a") as f:
                f.write(f"{ts}  {status}\n")

        return status

    except Exception as exc:
        error_msg = str(exc)[:120]
        row = [
            ticker, sector, cutoff_str,
            "error", 0, 0, 0,
            "", "", "", "", "", "", "", "", "", "", "", "", "",
            "", "", "", "", "", "",
            "", "", "ERROR", "", "", "", "", error_msg, ts,
        ]
        _write_row(excel_path, row, ticker, cutoff_str)
        status = f"✗ {ticker:<6} ERROR: {error_msg[:60]}"

        with _excel_lock:
            with open(log_path, "a") as f:
                f.write(f"{ts}  {status}\n")

        return status


# ── Main ──────────────────────────────────────────────────────────────────────

def _last_weekday(d: date) -> date:
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Multi-threaded backtest runner — outputs to Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/run_backtest_excel.py
  python scripts/run_backtest_excel.py --date 2025-12-31 --workers 4
  python scripts/run_backtest_excel.py --date 2025-12-31 --workers 6 --target-days 20
  python scripts/run_backtest_excel.py --date 2025-12-31 --workers 4 --tickers-per-sector 10
  python scripts/run_backtest_excel.py --date 2025-12-31 --sector "Energy" --workers 3
        """,
    )
    parser.add_argument(
        "--date", default=None,
        help="Cutoff date YYYY-MM-DD (default: last weekday before today)",
    )
    parser.add_argument(
        "--workers", type=int, default=4,
        help="Thread pool size. Recommended 3–6. More workers = faster but higher API rate risk.",
    )
    parser.add_argument(
        "--target-days", type=int, default=20,
        help="Max trade window in trading days (default: 20).",
    )
    parser.add_argument(
        "--tickers-per-sector", type=int, default=9,
        help="Max tickers sampled per sector (default: 9 → ~100 total across ~11 sectors).",
    )
    parser.add_argument(
        "--sector", default=None,
        help="Run only this sector (optional). Must match exactly, e.g. 'Energy'.",
    )
    args = parser.parse_args()

    # Resolve cutoff date
    if args.date:
        cutoff = datetime.strptime(args.date, "%Y-%m-%d").date()
    else:
        cutoff = _last_weekday(date.today() - timedelta(days=1))
    cutoff_str = cutoff.isoformat()

    # Output paths — folder named by cutoff date (not today — avoids confusion)
    run_date_str = date.today().isoformat()
    out_dir = OUTPUT_BASE / run_date_str
    out_dir.mkdir(parents=True, exist_ok=True)
    excel_path = out_dir / f"{run_date_str}.xlsx"
    log_path   = out_dir / "run_log.txt"

    # Select tickers
    tickers = _select_tickers(args.tickers_per_sector, args.sector)
    total = len(tickers)

    # Header banner
    print("=" * 72)
    print(f"  Backtest Excel Runner")
    print(f"  Cutoff : {cutoff_str}")
    print(f"  Window : {args.target_days} trading days")
    print(f"  Tickers: {total} ({args.tickers_per_sector}/sector{' — ' + args.sector if args.sector else ''})")
    print(f"  Workers: {args.workers}")
    print(f"  Output : {excel_path}")
    print("=" * 72)

    with open(log_path, "a") as f:
        f.write(f"\n{'='*60}\n")
        f.write(f"Run started: {datetime.now().isoformat()}\n")
        f.write(f"Cutoff: {cutoff_str}  |  Workers: {args.workers}  |  Tickers: {total}\n")
        f.write(f"{'='*60}\n")

    t0 = time.time()
    completed = 0
    errors    = 0
    trades    = 0

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {
            pool.submit(
                _predict_one,
                ticker, sector, cutoff_str,
                args.target_days, excel_path, log_path,
            ): (ticker, sector)
            for ticker, sector in tickers
        }

        for future in as_completed(futures):
            ticker, sector = futures[future]
            try:
                status = future.result()
                completed += 1
                if status.startswith("✓"):
                    trades += 1
                elif status.startswith("✗"):
                    errors += 1
                pct = completed / total * 100
                print(f"  [{completed:>3}/{total}  {pct:5.1f}%]  {status}")
            except Exception as exc:
                errors += 1
                completed += 1
                print(f"  [{completed:>3}/{total}] ✗ {ticker} unhandled: {exc}")

    elapsed = time.time() - t0
    print()
    print("=" * 72)
    print(f"  Done in {elapsed:.1f}s  |  Trades: {trades}  |  Errors: {errors}")
    print(f"  Excel : {excel_path}")
    print("=" * 72)

    with open(log_path, "a") as f:
        f.write(f"Run finished: {datetime.now().isoformat()}  elapsed={elapsed:.1f}s\n")
        f.write(f"Trades: {trades}  Errors: {errors}  Total: {total}\n")


if __name__ == "__main__":
    main()
