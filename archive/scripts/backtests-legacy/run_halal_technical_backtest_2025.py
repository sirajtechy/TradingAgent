#!/usr/bin/env python3
"""
run_halal_technical_backtest_2025.py
=====================================
Halal universe technical-agent backtest — January 2025 → December 2025.

Inputs
------
  • data/halal_universe/halal_master.json  — 1 236 Shariah-screened tickers
  • Polygon.io API key (.env)              — primary data source

Ticker selection
----------------
  • 10 tickers per sector  (4 mega/large + 3 mid + 3 small/micro/nano)
  • 12 sectors from halal_master.json  (N/A sector excluded)
  • Sectors with < 10 stocks: all available tickers used

Monthly windows
---------------
  • signal_date = 1st trading day of each month  (Jan → Dec 2025)
  • result_date = last calendar day of each month
  • 12 periods per ticker

Trade setup (per period, non-neutral signals only)
---------------------------------------------------
  • entry_date         = signal_date
  • exit_date_est      = signal_date + (holding_days_est × 1.4 calendar days)
  • entry_price        = Polygon close at signal_date
  • target_price       = entry ± 2.0 × ATR_14   [± based on direction]
  • stop_loss          = entry ∓ 1.5 × ATR_14
  • expected_profit_pct= |target − entry| / entry × 100
  • confidence_score   = experimental_score  (0 – 100, from rules engine)
  • profit_probability = score-mapped to 40 – 85 %
  • direction          = BULLISH | BEARISH  (NEUTRAL = skipped)
  • trade_duration_days= ADX-adjusted estimate  (filter: > 2 days only)

Exports
-------
  • JSON : backtest_output/halal_technical_2025/<TKR>_halal_tech_2025.json
  • Master: backtest_output/halal_technical_2025/master_halal_tech_2025.json
  • Excel : backtest_output/halal_technical_2025/halal_technical_backtest_2025.xlsx
              sheets: Summary | All Trades | <sector> × 12

Usage
-----
  # Full run (all sectors, 10 tickers each):
  python scripts/backtests/run_halal_technical_backtest_2025.py

  # Single sector:
  python scripts/backtests/run_halal_technical_backtest_2025.py --sector "Energy"

  # Override tickers:
  python scripts/backtests/run_halal_technical_backtest_2025.py --tickers AAPL,MSFT,NVDA

  # Resume (skip tickers already saved):
  python scripts/backtests/run_halal_technical_backtest_2025.py --resume

  # Quick smoke-test (1 ticker, 2 months):
  python scripts/backtests/run_halal_technical_backtest_2025.py --tickers AAPL --months 2
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# ── project root → sys.path ──────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

# Load .env before importing agent code (picks up POLYGON_API_KEY etc.)
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass  # dotenv not mandatory if env vars are already set

from agents.technical.service import analyze_ticker
from agents.technical.data_client import PolygonTechnicalClient
import paths


# ─────────────────────────────────────────────────────────────────────────────
# Paths & constants
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_DIR       = paths.HALAL_TECH_2025
HALAL_MASTER     = paths.HALAL_MASTER

# ATR multipliers for target and stop
ATR_TARGET_MULT = 2.0
ATR_STOP_MULT   = 1.5

# Score band → direction mapping (mirrors backtest.py)
BAND_TO_SIGNAL: Dict[str, str] = {
    "strong":         "BULLISH",
    "good":           "BULLISH",
    "mixed_positive": "BULLISH",
    "mixed":          "NEUTRAL",
    "weak":           "BEARISH",
}

# January – December 2025  (signal_date → result_date)
MONTHS_2025: List[Tuple[date, date]] = [
    (date(2025,  1,  1), date(2025,  1, 31)),
    (date(2025,  2,  1), date(2025,  2, 28)),
    (date(2025,  3,  1), date(2025,  3, 31)),
    (date(2025,  4,  1), date(2025,  4, 30)),
    (date(2025,  5,  1), date(2025,  5, 31)),
    (date(2025,  6,  1), date(2025,  6, 30)),
    (date(2025,  7,  1), date(2025,  7, 31)),
    (date(2025,  8,  1), date(2025,  8, 31)),
    (date(2025,  9,  1), date(2025,  9, 30)),
    (date(2025, 10,  1), date(2025, 10, 31)),
    (date(2025, 11,  1), date(2025, 11, 30)),
    (date(2025, 12,  1), date(2025, 12, 31)),
]

# Singleton Polygon client (shared across threads)
_price_client = PolygonTechnicalClient()


# ─────────────────────────────────────────────────────────────────────────────
# Ticker universe loader
# ─────────────────────────────────────────────────────────────────────────────

_LARGE_TIERS = {"mega", "large"}
_MID_TIERS   = {"mid"}
_SMALL_TIERS = {"small", "micro", "nano"}


def _load_ticker_universe(per_sector: int = 10) -> Dict[str, List[Dict]]:
    """
    Load halal_master.json and return {sector: [stock_meta, ...]} with
    ``per_sector`` tickers per entry, balanced across cap tiers:

        4 mega/large  +  3 mid  +  3 small/micro/nano  = 10

    Sectors with fewer than ``per_sector`` stocks get all available tickers.
    N/A sector is excluded.
    """
    with open(HALAL_MASTER) as fh:
        master = json.load(fh)

    result: Dict[str, List[Dict]] = {}

    for sector, stocks in master.get("stocks_by_sector", {}).items():
        if sector == "N/A":
            continue

        large = [s for s in stocks if s.get("market_cap_tier") in _LARGE_TIERS]
        mid   = [s for s in stocks if s.get("market_cap_tier") in _MID_TIERS]
        small = [s for s in stocks if s.get("market_cap_tier") in _SMALL_TIERS]

        n_large = min(4, len(large))
        n_mid   = min(3, len(mid))
        n_small = per_sector - n_large - n_mid

        selected: List[Dict] = (
            large[:n_large]
            + mid[:n_mid]
            + small[:max(0, n_small)]
        )

        # Pad to per_sector if still short (happens in small sectors)
        ticker_set = {s["ticker"] for s in selected}
        for s in large + mid + small:
            if len(selected) >= per_sector:
                break
            if s["ticker"] not in ticker_set:
                selected.append(s)
                ticker_set.add(s["ticker"])

        if selected:
            # Normalize market_cap_tier label for display
            for s in selected:
                tier = s.get("market_cap_tier", "unknown")
                if tier in _LARGE_TIERS:
                    s["cap_display"] = "large"
                elif tier in _MID_TIERS:
                    s["cap_display"] = "mid"
                else:
                    s["cap_display"] = "small"

            result[sector] = selected

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Trade setup calculator
# ─────────────────────────────────────────────────────────────────────────────

def _compute_trade_setup(
    result: Dict[str, Any],
    signal_date: date,
    result_date: date,
) -> Optional[Dict[str, Any]]:
    """
    Derive an ATR-based trade setup from a technical evaluation result.

    Returns None when:
      - signal is NEUTRAL or unknown
      - holding_days_est ≤ 2  (trade-duration filter)
      - required fields are absent
    """
    # Entry price
    entry_price: Optional[float] = result.get("as_of_price", {}).get("price")
    if not entry_price or entry_price <= 0:
        return None

    # Score & band
    exp  = result.get("experimental_score", {})
    score: Optional[float] = exp.get("score") if isinstance(exp, dict) and exp.get("available") else None
    band:  Optional[str]   = exp.get("band")  if isinstance(exp, dict) and exp.get("available") else None
    if score is None or band is None:
        return None

    signal = BAND_TO_SIGNAL.get(band, "NEUTRAL")
    if signal == "NEUTRAL":
        return None

    # Indicators
    key_ind: Dict[str, Any] = result.get("key_indicators", {}) or {}
    atr: float = key_ind.get("atr_14") or (entry_price * 0.02)   # fallback: 2 % of price
    adx: Optional[float] = key_ind.get("adx")
    rsi: Optional[float] = key_ind.get("rsi_14")

    # Target & stop prices
    if signal == "BULLISH":
        target_price = round(entry_price + ATR_TARGET_MULT * atr, 2)
        stop_loss    = round(entry_price - ATR_STOP_MULT  * atr, 2)
    else:  # BEARISH
        target_price = round(entry_price - ATR_TARGET_MULT * atr, 2)
        stop_loss    = round(entry_price + ATR_STOP_MULT  * atr, 2)

    expected_profit_pct = round(abs(target_price - entry_price) / entry_price * 100.0, 2)
    risk_pct            = round(abs(stop_loss    - entry_price) / entry_price * 100.0, 2)
    rr_ratio            = round(expected_profit_pct / risk_pct, 2) if risk_pct > 0 else None

    # Holding days estimate
    # ADX ≥ 30 → trending fast → 3–5 days
    # ADX 20–30 → moderate trend → 5–8 days
    # ADX < 20  → ranging → 8–15 days  (or ADX unknown)
    if adx is not None and adx >= 30:
        holding_days_est = max(3, round(score / 20))      # score=80 → 4 dys
    elif adx is not None and adx >= 20:
        holding_days_est = max(5, round(score / 12))      # score=72 → 6 dys
    else:
        holding_days_est = max(8, round(score / 8))       # score=72 → 9 dys

    # Trade-duration filter
    if holding_days_est <= 2:
        return None

    # Exit date (calendar days = business days × 1.4 covers weekends)
    exit_date_est = signal_date + timedelta(days=int(holding_days_est * 1.4))
    if exit_date_est > result_date:
        exit_date_est = result_date

    # Profit probability  (linear mapping → 40–85 %)
    # BULLISH: high score ↔ higher P(price goes up)
    # BEARISH: low  score ↔ higher P(price goes down)
    if signal == "BULLISH":
        profit_prob = round(min(85.0, max(40.0, score * 0.5 + 20.0)), 1)
    else:
        profit_prob = round(min(85.0, max(40.0, (100.0 - score) * 0.5 + 20.0)), 1)

    return {
        "entry_date":           signal_date.isoformat(),
        "exit_date_est":        exit_date_est.isoformat(),
        "result_date":          result_date.isoformat(),
        "entry_price":          round(entry_price, 2),
        "target_price":         target_price,
        "stop_loss":            stop_loss,
        "expected_profit_pct":  expected_profit_pct,
        "risk_pct":             risk_pct,
        "reward_risk_ratio":    rr_ratio,
        "confidence_score":     round(score, 1),
        "profit_probability":   profit_prob,
        "direction":            signal,
        "trade_duration_days":  holding_days_est,
        "score_band":           band,
        "atr_at_entry":         round(atr, 2),
        "rsi_at_entry":         rsi,
        "adx_at_entry":         adx,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Period runner (single month for one ticker)
# ─────────────────────────────────────────────────────────────────────────────

def _run_period(ticker: str, signal_date: date, result_date: date) -> Dict[str, Any]:
    """Evaluate one (signal_date → result_date) window for *ticker*."""

    # ── Step 1: run technical analysis at signal_date ──────────────────────
    try:
        tech_result = analyze_ticker(
            ticker=ticker,
            as_of_date=signal_date.isoformat(),
        )
    except Exception as exc:
        return {
            "month":          signal_date.strftime("%B %Y"),
            "signal_date":    signal_date.isoformat(),
            "result_date":    result_date.isoformat(),
            "error":          str(exc),
            "trade_setup":    None,
            "actual_outcome": None,
        }

    # ── Step 2: compute trade setup ────────────────────────────────────────
    trade_setup = _compute_trade_setup(tech_result, signal_date, result_date)

    # ── Step 3: actual period outcome ──────────────────────────────────────
    actual_outcome: Optional[Dict[str, Any]] = None
    try:
        end_bar    = _price_client.get_price_as_of(ticker, result_date)
        end_price  = end_bar.close
        start_price = tech_result["as_of_price"]["price"]
        return_pct  = round((end_price - start_price) / start_price * 100.0, 2)
        actual_dir  = "UP" if return_pct >= 0 else "DOWN"

        signal = trade_setup["direction"] if trade_setup else None
        if signal == "BULLISH":
            signal_correct: Optional[bool] = actual_dir == "UP"
        elif signal == "BEARISH":
            signal_correct = actual_dir == "DOWN"
        else:
            signal_correct = None

        actual_outcome = {
            "end_price":        round(end_price, 2),
            "end_price_date":   end_bar.bar_date.isoformat(),
            "price_return_pct": return_pct,
            "actual_direction": actual_dir,
            "signal_correct":   signal_correct,
        }
    except Exception as exc:
        actual_outcome = {"error": str(exc)}

    return {
        "month":          signal_date.strftime("%B %Y"),
        "signal_date":    signal_date.isoformat(),
        "result_date":    result_date.isoformat(),
        "trade_setup":    trade_setup,
        "actual_outcome": actual_outcome,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Ticker backtest runner (all 12 months)
# ─────────────────────────────────────────────────────────────────────────────

def _run_ticker(
    ticker: str,
    sector: str,
    cap_tier: str,
    months: List[Tuple[date, date]],
) -> Dict[str, Any]:
    """Run all monthly periods for one ticker and build the result dict."""
    periods: List[Dict[str, Any]] = []
    for signal_date, result_date in months:
        period = _run_period(ticker, signal_date, result_date)
        periods.append(period)
        time.sleep(0.15)  # light rate-limit buffer

    # Summary stats over directional (non-neutral) periods
    directional = [
        p for p in periods
        if p.get("actual_outcome")
        and isinstance((p.get("actual_outcome") or {}).get("signal_correct"), bool)
    ]
    correct  = sum(1 for p in directional if p["actual_outcome"]["signal_correct"])
    n_dir    = len(directional)
    accuracy = round(correct / n_dir * 100.0, 1) if n_dir > 0 else None

    setups = [p["trade_setup"] for p in periods if p.get("trade_setup")]
    def _avg(key: str) -> Optional[float]:
        vals = [s[key] for s in setups if s.get(key) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    return {
        "ticker":          ticker.upper(),
        "sector":          sector,
        "market_cap_tier": cap_tier,
        "periods":         periods,
        "summary": {
            "total_months":             len(periods),
            "directional_signals":      n_dir,
            "correct_signals":          correct,
            "accuracy_pct":             accuracy,
            "avg_expected_profit_pct":  _avg("expected_profit_pct"),
            "avg_confidence_score":     _avg("confidence_score"),
            "avg_profit_probability":   _avg("profit_probability"),
            "avg_trade_duration_days":  _avg("trade_duration_days"),
        },
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

TRADE_HEADERS = [
    "Ticker", "Sector", "Cap Tier", "Month",
    "Entry Date", "Exit Date Est",
    "Direction", "Entry Price", "Target Price", "Stop Loss",
    "Expected Profit %", "Risk %", "R:R Ratio",
    "Confidence Score", "Profit Probability %", "Trade Duration (Days)",
    "ATR at Entry", "RSI at Entry", "ADX at Entry",
    "Actual End Price", "Actual Return %", "Actual Direction", "Signal Correct",
]

SUMMARY_HEADERS = [
    "Sector", "Ticker", "Cap Tier",
    "Directional Signals", "Correct Signals", "Accuracy %",
    "Avg Confidence", "Avg Profit Prob %",
    "Avg Expected Profit %", "Avg Trade Duration (Days)",
]


def _build_trade_row(ticker_data: Dict, period: Dict) -> Optional[Dict]:
    """Build a flat row dict from a period dict, or None if no trade setup."""
    ts = period.get("trade_setup")
    if ts is None:
        return None
    ao = period.get("actual_outcome") or {}
    return {
        "Ticker":                  ticker_data["ticker"],
        "Sector":                  ticker_data["sector"],
        "Cap Tier":                ticker_data["market_cap_tier"],
        "Month":                   period["month"],
        "Entry Date":              ts.get("entry_date"),
        "Exit Date Est":           ts.get("exit_date_est"),
        "Direction":               ts.get("direction"),
        "Entry Price":             ts.get("entry_price"),
        "Target Price":            ts.get("target_price"),
        "Stop Loss":               ts.get("stop_loss"),
        "Expected Profit %":       ts.get("expected_profit_pct"),
        "Risk %":                  ts.get("risk_pct"),
        "R:R Ratio":               ts.get("reward_risk_ratio"),
        "Confidence Score":        ts.get("confidence_score"),
        "Profit Probability %":    ts.get("profit_probability"),
        "Trade Duration (Days)":   ts.get("trade_duration_days"),
        "ATR at Entry":            ts.get("atr_at_entry"),
        "RSI at Entry":            ts.get("rsi_at_entry"),
        "ADX at Entry":            ts.get("adx_at_entry"),
        "Actual End Price":        ao.get("end_price"),
        "Actual Return %":         ao.get("price_return_pct"),
        "Actual Direction":        ao.get("actual_direction"),
        "Signal Correct":          ao.get("signal_correct"),
    }


def _export_excel(all_results: List[Dict], output_path: Path) -> None:
    """Write a multi-sheet Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠  openpyxl not installed — skipping Excel export.")
        print("     Run:  pip install openpyxl")
        return

    # Colour palettes
    FILL_HEADER_GREEN  = PatternFill("solid", fgColor="1E8449")
    FILL_HEADER_BLUE   = PatternFill("solid", fgColor="154360")
    FILL_HEADER_ORANGE = PatternFill("solid", fgColor="784212")
    FONT_HEADER        = Font(bold=True, color="FFFFFF", size=10)
    FILL_CORRECT       = PatternFill("solid", fgColor="A9DFBF")   # green
    FILL_WRONG         = PatternFill("solid", fgColor="F5B7B1")   # red
    FILL_BULLISH       = PatternFill("solid", fgColor="EAF2FF")   # light blue
    FILL_BEARISH       = PatternFill("solid", fgColor="FDF2F8")   # light pink
    ALIGN_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ── Build flat row lists ───────────────────────────────────────────────
    flat_rows:    List[Dict] = []
    sector_rows:  Dict[str, List[Dict]] = {}

    for td in all_results:
        for p in td["periods"]:
            row = _build_trade_row(td, p)
            if row is None:
                continue
            flat_rows.append(row)
            sector_rows.setdefault(td["sector"], []).append(row)

    def _write_trade_sheet(ws, rows: List[Dict], fill: PatternFill) -> None:
        # Header row
        for ci, header in enumerate(TRADE_HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.font  = FONT_HEADER
            cell.fill  = fill
            cell.alignment = ALIGN_CENTER

        # Data rows
        for ri, row in enumerate(rows, 2):
            correct   = row.get("Signal Correct")
            direction = row.get("Direction", "")
            for ci, header in enumerate(TRADE_HEADERS, 1):
                cell = ws.cell(row=ri, column=ci, value=row.get(header))
                cell.alignment = ALIGN_CENTER
                # Row-level colouring
                if correct is True:
                    cell.fill = FILL_CORRECT
                elif correct is False:
                    cell.fill = FILL_WRONG
                elif direction == "BULLISH":
                    cell.fill = FILL_BULLISH
                elif direction == "BEARISH":
                    cell.fill = FILL_BEARISH

        # Column widths
        for ci in range(1, len(TRADE_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16
        ws.freeze_panes = "A2"

    # ── Workbook ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # --- Summary sheet -------------------------------------------------------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    for ci, header in enumerate(SUMMARY_HEADERS, 1):
        cell = ws_sum.cell(row=1, column=ci, value=header)
        cell.font   = FONT_HEADER
        cell.fill   = FILL_HEADER_ORANGE
        cell.alignment = ALIGN_CENTER

    ri = 2
    for td in sorted(all_results, key=lambda x: (x["sector"], x["ticker"])):
        s   = td["summary"]
        acc = s.get("accuracy_pct")
        row_vals = [
            td["sector"],
            td["ticker"],
            td["market_cap_tier"],
            s.get("directional_signals"),
            s.get("correct_signals"),
            f"{acc:.1f}%" if acc is not None else "N/A",
            s.get("avg_confidence_score"),
            s.get("avg_profit_probability"),
            s.get("avg_expected_profit_pct"),
            s.get("avg_trade_duration_days"),
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            cell.alignment = ALIGN_CENTER
            if ci == 6 and acc is not None:
                cell.fill = FILL_CORRECT if acc >= 60 else (FILL_WRONG if acc < 45 else None) or PatternFill()
        ri += 1

    for ci in range(1, len(SUMMARY_HEADERS) + 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 20
    ws_sum.freeze_panes = "A2"

    # --- All Trades sheet ----------------------------------------------------
    ws_all = wb.create_sheet("All Trades")
    _write_trade_sheet(ws_all, flat_rows, FILL_HEADER_GREEN)

    # --- Per-sector sheets ---------------------------------------------------
    for sector, rows in sorted(sector_rows.items()):
        sheet_name = sector[:31]   # Excel max 31 chars
        ws_sec = wb.create_sheet(sheet_name)
        _write_trade_sheet(ws_sec, rows, FILL_HEADER_BLUE)

    wb.save(output_path)
    print(f"  ✓ Excel  → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI helpers
# ─────────────────────────────────────────────────────────────────────────────

def _print_header(n_sectors: int, n_tickers: int, n_months: int, workers: int) -> None:
    print()
    print("█" * 68)
    print("█  HALAL TECHNICAL BACKTEST  —  Jan 2025 → Dec 2025")
    print("█" * 68)
    print(f"  Sectors    : {n_sectors}")
    print(f"  Tickers    : {n_tickers}")
    print(f"  Months     : {n_months}  (Jan – Dec 2025)")
    print(f"  Data source: Polygon.io  (yfinance fallback)")
    print(f"  Filter     : trade_duration_days > 2")
    print(f"  Workers    : {workers}")
    print(f"  Output dir : {OUTPUT_DIR}")
    print()


def _print_sector_summary(all_results: List[Dict], sector_meta: Dict) -> None:
    print()
    print("─" * 68)
    print(f"  {'SECTOR':<28}  {'TICKERS':>7}  {'TRADES':>7}  {'ACCURACY':>10}")
    print("─" * 68)
    per_sector: Dict[str, Dict] = {}
    for td in all_results:
        sec = td["sector"]
        if sec not in per_sector:
            per_sector[sec] = {"n": 0, "dir": 0, "cor": 0}
        per_sector[sec]["n"]   += 1
        per_sector[sec]["dir"] += td["summary"]["directional_signals"] or 0
        per_sector[sec]["cor"] += td["summary"]["correct_signals"]     or 0

    total_dir = total_cor = total_tickers = 0
    for sec in sorted(per_sector):
        d = per_sector[sec]
        acc = f"{d['cor'] / d['dir'] * 100:.1f}%" if d["dir"] > 0 else "N/A"
        print(f"  {sec:<28}  {d['n']:>7}  {d['dir']:>7}  {acc:>10}")
        total_dir     += d["dir"]
        total_cor     += d["cor"]
        total_tickers += d["n"]
    print("─" * 68)
    overall = f"{total_cor / total_dir * 100:.1f}%" if total_dir > 0 else "N/A"
    print(f"  {'OVERALL':<28}  {total_tickers:>7}  {total_dir:>7}  {overall:>10}")
    print("─" * 68)
    print()


# ─────────────────────────────────────────────────────────────────────────────
# Master JSON builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_master(
    all_results: List[Dict],
    months: List[Tuple[date, date]],
) -> Dict[str, Any]:
    """Aggregate all ticker results into a single master JSON."""

    all_trades: List[Dict] = []
    sector_agg: Dict[str, Dict] = {}

    for td in all_results:
        sec = td["sector"]
        if sec not in sector_agg:
            sector_agg[sec] = {
                "tickers":   [],
                "total_directional": 0,
                "total_correct":     0,
            }
        sector_agg[sec]["tickers"].append(td["ticker"])
        sector_agg[sec]["total_directional"] += td["summary"]["directional_signals"] or 0
        sector_agg[sec]["total_correct"]     += td["summary"]["correct_signals"]     or 0

        for p in td["periods"]:
            row = _build_trade_row(td, p)
            if row:
                all_trades.append(row)

    # Sector-level accuracy
    for sec, agg in sector_agg.items():
        n = agg["total_directional"]
        c = agg["total_correct"]
        agg["accuracy_pct"] = round(c / n * 100.0, 1) if n > 0 else None

    return {
        "metadata": {
            "backtest_window":       "January 2025 – December 2025",
            "months":                [m[0].isoformat() for m in months],
            "total_tickers":         len(all_results),
            "sectors":               sorted(sector_agg.keys()),
            "total_trades":          len(all_trades),
            "data_source":           "Polygon.io (primary) / yfinance (fallback)",
            "trade_duration_filter": "holding_days_est > 2",
            "atr_target_mult":       ATR_TARGET_MULT,
            "atr_stop_mult":         ATR_STOP_MULT,
            "generated_at":          date.today().isoformat(),
        },
        "sector_summary":   sector_agg,
        "all_trades":       all_trades,
        "ticker_summaries": {td["ticker"]: td["summary"] for td in all_results},
    }


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Halal universe technical backtest — 2025",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--sector",     default=None,   help="Run only this sector")
    parser.add_argument("--tickers",    default=None,   help="Comma-separated ticker overrides")
    parser.add_argument("--per-sector", type=int, default=10, help="Tickers per sector (default: 10)")
    parser.add_argument("--workers",    type=int, default=4,  help="Concurrent threads (default: 4)")
    parser.add_argument("--months",     type=int, default=12, help="Months to run 1-12 (default: 12)")
    parser.add_argument("--resume",     action="store_true",  help="Skip tickers already saved")
    parser.add_argument("--no-excel",   action="store_true",  help="Skip Excel export")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Choose months window ──────────────────────────────────────────────
    months = MONTHS_2025[: max(1, min(args.months, 12))]

    # ── Build ticker universe ─────────────────────────────────────────────
    if args.tickers:
        # Manual ticker override — assign to 'Custom' sector
        ticker_meta = [
            {
                "ticker":          t.strip().upper(),
                "name":            t.strip().upper(),
                "market_cap_tier": "large",
                "cap_display":     "large",
            }
            for t in args.tickers.split(",")
            if t.strip()
        ]
        universe: Dict[str, List[Dict]] = {"Custom": ticker_meta}
    else:
        universe = _load_ticker_universe(per_sector=args.per_sector)

    # Optionally filter to one sector
    if args.sector:
        universe = {
            k: v for k, v in universe.items()
            if k.lower() == args.sector.lower()
        }
        if not universe:
            print(f"  ERROR: sector '{args.sector}' not found in universe.")
            sys.exit(1)

    total_tickers = sum(len(v) for v in universe.values())
    _print_header(len(universe), total_tickers, len(months), args.workers)

    # ── Build task list (skip already cached if --resume) ─────────────────
    tasks: List[Tuple[str, str, str]] = []   # (ticker, sector, cap_tier)
    all_results: List[Dict] = []

    # Load cached results first
    if args.resume:
        for cache_path in sorted(OUTPUT_DIR.glob("*_halal_tech_2025.json")):
            if cache_path.stem.startswith("master"):
                continue  # skip the master JSON — it has a different structure
            try:
                with open(cache_path) as fh:
                    data = json.load(fh)
                    if "ticker" in data:
                        all_results.append(data)
            except (json.JSONDecodeError, KeyError):
                pass
        cached_tickers = {td["ticker"] for td in all_results}
        print(f"  Loaded {len(all_results)} cached results.", flush=True)
    else:
        cached_tickers: set = set()

    for sector, stocks in universe.items():
        for stock in stocks:
            ticker   = stock["ticker"].upper()
            cap_tier = stock.get("cap_display", stock.get("market_cap_tier", "unknown"))
            if args.resume and ticker in cached_tickers:
                print(f"    [{ticker:<8}] ← cache hit", flush=True)
                continue
            tasks.append((ticker, sector, cap_tier))

    if not tasks and not all_results:
        print("  Nothing to run. Exiting.")
        sys.exit(0)

    # ── Run backtest with thread pool ─────────────────────────────────────
    if tasks:
        print(f"\n  Running {len(tasks)} tickers × {len(months)} months ...\n")

    def _worker(task: Tuple[str, str, str]) -> Dict:
        ticker, sector, cap_tier = task
        print(f"  → {ticker:<8}  [{sector}] [{cap_tier}]", flush=True)
        t0   = time.time()
        data = _run_ticker(ticker, sector, cap_tier, months)
        elapsed = time.time() - t0
        s   = data["summary"]
        acc_str = f"{s['accuracy_pct']:.0f}%" if s.get("accuracy_pct") is not None else "N/A"
        print(
            f"  ✓ {ticker:<8}  dir={s['directional_signals']}/{len(months)}"
            f"  acc={acc_str}  ({elapsed:.0f}s)",
            flush=True,
        )
        return data

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {pool.submit(_worker, task): task for task in tasks}
        for fut in as_completed(futures):
            ticker_data = fut.result()
            all_results.append(ticker_data)

            # Save per-ticker JSON immediately to disk
            cache_path = OUTPUT_DIR / f"{ticker_data['ticker']}_halal_tech_2025.json"
            with open(cache_path, "w") as fh:
                json.dump(ticker_data, fh, indent=2, default=str)

    # ── Write master JSON ─────────────────────────────────────────────────
    master = _build_master(all_results, months)
    master_path = OUTPUT_DIR / "master_halal_tech_2025.json"
    with open(master_path, "w") as fh:
        json.dump(master, fh, indent=2, default=str)
    print(f"\n  ✓ Master JSON → {master_path}")

    # ── Excel export ──────────────────────────────────────────────────────
    if not args.no_excel:
        xlsx_path = OUTPUT_DIR / "halal_technical_backtest_2025.xlsx"
        print("  Building Excel workbook ...")
        _export_excel(all_results, xlsx_path)

    # ── Print summary table ───────────────────────────────────────────────
    _print_sector_summary(all_results, universe)
    print(f"  Output → {OUTPUT_DIR}\n")


if __name__ == "__main__":
    main()
