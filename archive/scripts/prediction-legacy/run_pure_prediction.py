#!/usr/bin/env python3
"""
run_pure_prediction.py — Pure prediction backtest (NO forward-looking)

Generates model predictions at multiple cutoff dates without ever looking
at future data. This is for MODEL EVALUATION, not trade simulation.

Design Philosophy
─────────────────
• Uses ONLY data ≤ cutoff date for prediction
• Predicts: BUY/HOLD/AVOID + confidence + targets
• NEVER looks forward to calculate outcomes
• Outcome evaluation done later by comparing predictions to actual results

This is NOT a backtest — it's a prediction archive for model validation.

Usage
─────
  # Single date prediction
  python scripts/run_pure_prediction.py --date 2025-09-30

  # Multiple dates (monthly for 2025)
  python scripts/run_pure_prediction.py --year 2025 --interval monthly

Parameters
──────────
  --date          Single cutoff date (YYYY-MM-DD)
  --year          Run for entire year (e.g., 2025)
  --start         Start date (YYYY-MM-DD)
  --end           End date (YYYY-MM-DD)
  --interval      weekly, biweekly, monthly (default: monthly)
  --workers       Thread pool size. Default: 4
  --tickers-per-sector  Max tickers per sector. Default: 9

Output
──────
  data/output/predictions/<YYYY-MM-DD>/
      predictions.json     — Pure predictions, NO outcomes
      run_log.txt          — Execution log
"""

from __future__ import annotations

import argparse
import json
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Project root ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

HALAL_JSON = ROOT / "data" / "halal_universe" / "halal_sector_tickers.json"
OUTPUT_BASE = ROOT / "data" / "output" / "predictions"

# ── Excel columns (PREDICTION ONLY - no outcomes) ────────────────────────────
HEADERS = [
    "Ticker", "Sector", "Cutoff Date", "Signal", "Confidence Score",
    "Tech Score", "Fund Score", 
    "Pattern Name", "Pattern Start", "Pattern End", "Breakout Date",
    "Entry Conviction", "Price Extension %",
    "Entry Price (Est.)", "Retest Zone Low", "Retest Zone High",
    "Target Price", "Stop Loss", "R/R Ratio",
    "ATR", "ADX", "RSI",
    "Est. Days to Target",
    "No Trade Reason",
    "Prediction Timestamp",
]

_excel_lock = threading.Lock()


def _ensure_workbook(path: Path) -> openpyxl.Workbook:
    """Load existing workbook or create new with headers."""
    if path.exists():
        return openpyxl.load_workbook(path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF203864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    return wb


def _find_existing_row(ws, ticker: str, cutoff_date: str) -> Optional[int]:
    """Return 1-based row index if ticker+cutoff exists."""
    for row_idx in range(2, ws.max_row + 1):
        if (ws.cell(row_idx, 1).value == ticker and
                str(ws.cell(row_idx, 3).value) == cutoff_date):
            return row_idx
    return None


def _write_row(path: Path, row_values: List[Any], ticker: str, cutoff_date: str) -> None:
    """Thread-safe write to Excel."""
    with _excel_lock:
        wb = _ensure_workbook(path)
        ws = wb.active

        existing_row = _find_existing_row(ws, ticker, cutoff_date)
        if existing_row:
            for col_idx, val in enumerate(row_values, start=1):
                ws.cell(row=existing_row, column=col_idx).value = val
        else:
            ws.append(row_values)

        wb.save(path)


def _select_tickers(
    tickers_per_sector: int,
    sector_filter: Optional[str],
) -> List[Tuple[str, str]]:
    """Return list of (ticker, sector) tuples."""
    raw: Dict[str, List[str]] = json.loads(HALAL_JSON.read_text())
    seen: set = set()
    result: List[Tuple[str, str]] = []

    for sector, tickers in raw.items():
        if sector == "N/A":
            continue
        if sector_filter and sector.lower() != sector_filter.lower():
            continue
        count = 0
        for t in tickers:
            if " " in t or len(t) > 6:
                continue
            if t in seen:
                continue
            seen.add(t)
            result.append((t, sector))
            count += 1
            if count >= tickers_per_sector:
                break

    return result


def _predict_one_pure(
    ticker: str,
    sector: str,
    cutoff_str: str,
    excel_path: Path,
    log_path: Path,
) -> str:
    """
    Generate PURE PREDICTION for one ticker.
    
    ✅ Uses ONLY data ≤ cutoff_date
    ❌ NEVER looks forward to calculate outcomes
    """
    from agents.orchestrator.service import predict
    
    ts = datetime.now().isoformat(timespec="seconds")

    try:
        # Call orchestrator with cutoff date
        # This uses ONLY data before cutoff for prediction
        result = predict(ticker=ticker, as_of_date=cutoff_str)
        
        sentiment = result.get("sentiment", "neutral")
        confidence_score = result.get("confidence_score", 0)
        tech_score = result.get("tech_score", 0)
        fund_score = result.get("fund_score", 0)
        
        # Determine signal
        if sentiment == "bullish" and confidence_score >= 60:
            signal = "BUY"
        elif sentiment == "bearish":
            signal = "AVOID"
        else:
            signal = "HOLD"
        
        # Get key indicators and patterns
        key_ind = result.get("tech", {}).get("key_indicators", {})
        patterns = result.get("tech", {}).get("patterns", [])
        
        # Best pattern (if any)
        pattern_name = ""
        pattern_start = ""
        pattern_end = ""
        breakout_date = ""
        entry_price = None
        target_price = None
        stop_loss = None
        rr_ratio = None
        retest_low = None
        retest_high = None
        entry_conviction = signal  # HOLD/AVOID for non-BUY
        price_extension = ""
        est_days = ""
        
        if patterns and signal == "BUY":
            best = patterns[0] if isinstance(patterns[0], dict) else {}
            pattern_name = best.get("name", "")
            pattern_start = best.get("start_date", "")
            pattern_end = best.get("end_date", "")
            breakout_date = best.get("breakout_date", "")
            
            # Entry price from last close
            current_price = key_ind.get("close")
            if current_price:
                entry_price = float(current_price)
                
                # Calculate targets based on ATR
                atr = key_ind.get("atr_14")
                if atr:
                    atr_f = float(atr)
                    stop_loss = round(entry_price - 2.0 * atr_f, 2)
                    
                    # Use pattern target if available
                    pattern_target = best.get("pattern_target") or best.get("target")
                    if pattern_target:
                        target_price = round(float(pattern_target), 2)
                    else:
                        target_price = round(entry_price + 3.0 * atr_f, 2)
                    
                    # R/R ratio
                    risk = entry_price - stop_loss
                    reward = target_price - entry_price
                    rr_ratio = round(reward / risk, 2) if risk > 0 else ""
                    
                    # Est days to target (rough estimate)
                    est_days = 15  # Default estimate
                    
                    # Retest zone
                    breakout_price = best.get("breakout_price")
                    if breakout_price:
                        retest_low = round(float(breakout_price), 2)
                        retest_high = round(float(breakout_price) + atr_f, 2)
                        
                        # Entry conviction based on price extension
                        extension = ((entry_price - float(breakout_price)) / float(breakout_price)) * 100
                        price_extension = round(extension, 2)
                        
                        if extension < 3:
                            entry_conviction = "IMMEDIATE"
                        elif extension < 8:
                            entry_conviction = "RETEST_ENTRY"
                        else:
                            entry_conviction = "WAIT_FOR_RETEST"
        
        # Build prediction row (NO OUTCOMES)
        row = [
            ticker,
            sector,
            cutoff_str,
            signal,
            round(confidence_score, 1),
            round(tech_score, 1),
            round(fund_score, 1),
            pattern_name,
            pattern_start,
            pattern_end,
            breakout_date,
            entry_conviction,
            price_extension,
            entry_price,
            retest_low,
            retest_high,
            target_price,
            stop_loss,
            rr_ratio,
            key_ind.get("atr_14"),
            key_ind.get("adx"),
            key_ind.get("rsi_14"),
            est_days,
            "" if signal == "BUY" else result.get("no_trade_reason", ""),
            ts,
        ]
        
        _write_row(excel_path, row, ticker, cutoff_str)
        
        # Log
        status = f"✓ {ticker:<6} [{signal:<6}] confidence={confidence_score:.1f} {entry_conviction if signal=='BUY' else ''}"
        with _excel_lock:
            with open(log_path, "a") as f:
                f.write(f"{ts}  {status}\n")
        
        return status

    except Exception as exc:
        error_msg = str(exc)[:120]
        row = [
            ticker, sector, cutoff_str,
            "ERROR", 0, 0, 0,
            "", "", "", "", "", "", "", "", "", "", "", "",
            "", "", "", "", error_msg, ts,
        ]
        _write_row(excel_path, row, ticker, cutoff_str)
        status = f"✗ {ticker:<6} ERROR: {error_msg[:60]}"
        
        with _excel_lock:
            with open(log_path, "a") as f:
                f.write(f"{ts}  {status}\n")
        
        return status


def _last_weekday(d: date) -> date:
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def generate_cutoff_dates(
    start_date: date,
    end_date: date,
    interval: str
) -> List[date]:
    """Generate cutoff dates based on interval."""
    dates = []
    current = start_date
    
    if interval == "weekly":
        delta = timedelta(days=7)
    elif interval == "biweekly":
        delta = timedelta(days=14)
    elif interval == "monthly":
        delta = timedelta(days=30)
    else:
        raise ValueError(f"Unknown interval: {interval}")
    
    while current <= end_date:
        weekday = _last_weekday(current)
        if weekday not in dates:
            dates.append(weekday)
        current += delta
    
    return sorted(dates)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Pure prediction generator (NO forward-looking)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--date", help="Single cutoff date YYYY-MM-DD")
    parser.add_argument("--year", type=int, help="Year to predict (e.g., 2025)")
    parser.add_argument("--start", help="Start date YYYY-MM-DD")
    parser.add_argument("--end", help="End date YYYY-MM-DD")
    parser.add_argument(
        "--interval",
        choices=["weekly", "biweekly", "monthly"],
        default="monthly",
        help="Date interval (default: monthly)"
    )
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--tickers-per-sector", type=int, default=9)
    parser.add_argument("--sector", default=None)
    
    args = parser.parse_args()
    
    # Determine cutoff dates
    if args.date:
        cutoff_dates = [datetime.strptime(args.date, "%Y-%m-%d").date()]
        label = args.date
    elif args.year:
        start = date(args.year, 1, 1)
        end = date(args.year, 12, 31)
        cutoff_dates = generate_cutoff_dates(start, end, args.interval)
        label = f"{args.year}_{args.interval}"
    elif args.start and args.end:
        start = datetime.strptime(args.start, "%Y-%m-%d").date()
        end = datetime.strptime(args.end, "%Y-%m-%d").date()
        cutoff_dates = generate_cutoff_dates(start, end, args.interval)
        label = f"{args.start}_to_{args.end}"
    else:
        print("ERROR: Specify --date OR --year OR --start/--end")
        sys.exit(1)
    
    # Output directory
    out_dir = OUTPUT_BASE / label
    out_dir.mkdir(parents=True, exist_ok=True)
    excel_path = out_dir / "predictions.xlsx"
    log_path = out_dir / "run_log.txt"
    
    # Select tickers
    tickers = _select_tickers(args.tickers_per_sector, args.sector)
    total = len(tickers)
    
    print("=" * 72)
    print(f"  Pure Prediction Generator (NO Forward-Looking)")
    print(f"  Cutoff dates: {len(cutoff_dates)}")
    for d in cutoff_dates:
        print(f"    - {d.isoformat()}")
    print(f"  Tickers: {total} ({args.tickers_per_sector}/sector)")
    print(f"  Workers: {args.workers}")
    print(f"  Output: {excel_path}")
    print("=" * 72)
    print()
    print("⚠️  IMPORTANT: NO OUTCOME CALCULATION")
    print("   • Uses ONLY data ≤ cutoff date")
    print("   • NEVER looks at future data")
    print("   • For model evaluation, not backtesting")
    print("=" * 72)
    
    with open(log_path, "a") as f:
        f.write(f"\n{'='*60}\n")
        f.write(f"Pure Prediction Run: {datetime.now().isoformat()}\n")
        f.write(f"Cutoff dates: {[d.isoformat() for d in cutoff_dates]}\n")
        f.write(f"Tickers: {total}\n")
        f.write(f"{'='*60}\n")
    
    # Run predictions for each cutoff date
    for cutoff_date in cutoff_dates:
        cutoff_str = cutoff_date.isoformat()
        print(f"\n{'='*72}")
        print(f"Generating predictions for: {cutoff_str}")
        print(f"{'='*72}")
        
        t0 = time.time()
        completed = 0
        errors = 0
        
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            futures = {
                pool.submit(
                    _predict_one_pure,
                    ticker, sector, cutoff_str,
                    excel_path, log_path,
                ): (ticker, sector)
                for ticker, sector in tickers
            }
            
            for future in as_completed(futures):
                ticker, sector = futures[future]
                try:
                    status = future.result()
                    completed += 1
                    if status.startswith("✗"):
                        errors += 1
                    pct = completed / total * 100
                    print(f"  [{completed:>3}/{total}  {pct:5.1f}%]  {status}")
                except Exception as exc:
                    errors += 1
                    completed += 1
                    print(f"  [{completed:>3}/{total}] ✗ {ticker} unhandled: {exc}")
        
        elapsed = time.time() - t0
        print(f"\nCompleted {cutoff_str} in {elapsed:.1f}s | Errors: {errors}")
    
    # Convert to JSON
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)
    
    json_path = out_dir / "predictions.json"
    with open(json_path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    
    print("\n" + "=" * 72)
    print(f"  PURE PREDICTIONS COMPLETE")
    print(f"  Total records: {len(data)}")
    print(f"  Excel: {excel_path}")
    print(f"  JSON: {json_path}")
    print("=" * 72)


if __name__ == "__main__":
    main()
