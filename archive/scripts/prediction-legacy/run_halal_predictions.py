#!/usr/bin/env python3
"""
run_halal_predictions.py — Batch predict_trade() runner for the full Halal universe.

Reads data/halal_universe/halal_sector_tickers.json and runs the complete
TA + FA orchestrator pipeline for every ticker as of today's date.

Usage
─────
    python scripts/run_halal_predictions.py
    python scripts/run_halal_predictions.py --workers 6
    python scripts/run_halal_predictions.py --date 2026-04-04
    python scripts/run_halal_predictions.py --sector "Information Technology"

Output
──────
    data/output/predictions/halal_<date>/         ← one JSON per ticker
    data/output/predictions/halal_<date>_summary.json
    data/output/predictions/halal_<date>_summary.md
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Ensure project root is on path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

HALAL_UNIVERSE = Path(__file__).parent.parent / "data" / "halal_universe" / "halal_sector_tickers.json"
OUTPUT_BASE    = Path(__file__).parent.parent / "data" / "output" / "predictions"

# ---------------------------------------------------------------------------
# 50-ticker sector universe (5 sectors × 10 tickers)
# ---------------------------------------------------------------------------
SECTORS: Dict[str, List[str]] = {
    "Technology": [
        "AAPL", "MSFT", "NVDA", "GOOGL", "META",
        "AMZN", "TSLA", "ORCL", "ANET", "CRM",
    ],
    "Healthcare": [
        "JNJ", "UNH", "LLY", "ABBV", "MRK",
        "PFE", "BMY", "CVS", "CI", "ABT",
    ],
    "Financials": [
        "JPM", "BAC", "WFC", "GS", "MS",
        "V", "MA", "AXP", "BLK", "C",
    ],
    "Consumer_Staples": [
        "PEP", "KO", "PG", "WMT", "COST",
        "MCD", "PM", "MO", "GIS", "CL",
    ],
    "Energy": [
        "XOM", "CVX", "COP", "SLB", "OXY",
        "PSX", "VLO", "MPC", "EOG", "HAL",
    ],
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _last_trading_day(d: date) -> date:
    """Return d if weekday, else step back to Friday."""
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def _load_tickers(
    sector_filter: Optional[str] = None,
    universe: str = "halal",
) -> Dict[str, List[str]]:
    """Load and deduplicate tickers.

    universe='halal'   → reads halal_sector_tickers.json (1 225 tickers)
    universe='sectors' → uses the hard-coded 50-ticker SECTORS dict
    """
    if universe == "sectors":
        if sector_filter:
            return {
                k: v for k, v in SECTORS.items()
                if k.lower() == sector_filter.lower()
            }
        return {k: list(v) for k, v in SECTORS.items()}

    # halal universe from JSON
    raw: Dict[str, List[str]] = json.loads(HALAL_UNIVERSE.read_text())
    seen = set()
    result: Dict[str, List[str]] = {}
    for sector, tickers in raw.items():
        if sector_filter and sector.lower() != sector_filter.lower():
            continue
        clean = []
        for t in tickers:
            if " " in t or len(t) > 6:
                continue   # skip WI tickers and malformed symbols
            if t not in seen:
                seen.add(t)
                clean.append(t)
        if clean:
            result[sector] = clean
    return result


# ---------------------------------------------------------------------------
# Worker
# ---------------------------------------------------------------------------

def _predict_ticker(
    ticker: str,
    sector: str,
    cutoff_str: str,
    target_days: int,
) -> Tuple[str, str, Optional[Dict[str, Any]]]:
    """
    Run predict_trade() for a single ticker.
    Returns (ticker, sector, result_dict | None).
    """
    from agents.technical.service import predict_trade
    try:
        result = predict_trade(ticker=ticker, cutoff_date=cutoff_str, target_days=target_days)
        result["sector"] = sector
        trade    = result.get("trade")
        outcome  = trade["exit_outcome"] if trade else "NO_TRADE"
        source   = trade["entry_source"][:40] if trade else (result.get("no_trade_reason") or "")[:40]
        signal   = result.get("sentiment", "?")
        score    = result.get("confidence_score", 0)
        print(f"  {'✓' if trade else '·'} {ticker:<6} [{signal:<8}] score={score:.0f}  {outcome:<12} {source}", flush=True)
        return ticker, sector, result
    except Exception as exc:
        msg = str(exc)
        print(f"  ✗ {ticker:<6} ERROR: {msg[:70]}", flush=True)
        return ticker, sector, {"ticker": ticker, "sector": sector, "error": msg}


# ---------------------------------------------------------------------------
# Summary builder
# ---------------------------------------------------------------------------

def _build_summary(
    results: List[Dict[str, Any]],
    cutoff_date: str,
    target_days: int,
    elapsed: float,
) -> Dict[str, Any]:
    trades, no_trades, errors = [], [], []

    for r in results:
        if "error" in r and "sentiment" not in r:
            errors.append(r)
            continue
        trade = r.get("trade")
        entry = {
            "ticker":           r.get("ticker", "?"),
            "sector":           r.get("sector", "Unknown"),
            "sentiment":        r.get("sentiment", "?"),
            "confidence_score": r.get("confidence_score", 0),
            "tech_score":       r.get("tech_score", 0),
            "fund_score":       r.get("fund_score", 0),
            "conflict":         r.get("conflict_detected", False),
            "no_trade_reason":  r.get("no_trade_reason"),
        }
        if trade:
            entry.update({
                "entry_date":          trade["entry_date"],
                "entry_price":         trade["entry_price"],
                "entry_price_note":    trade.get("entry_price_note", ""),
                "entry_source":        trade["entry_source"],
                "entry_earliest":      trade.get("entry_earliest", trade["entry_date"]),
                "entry_conviction":    trade.get("entry_conviction", ""),
                "price_extension_pct": trade.get("price_extension_pct"),
                "retest_zone_low":     trade.get("retest_zone_low"),
                "retest_zone_high":    trade.get("retest_zone_high"),
                "exit_date":           trade["exit_date"],
                "exit_price":          trade["exit_price"],
                "exit_outcome":        trade["exit_outcome"],
                "holding_days":        trade["holding_days"],
                "gross_profit_pct":    trade["gross_profit_pct"],
                "net_profit_pct":      trade["net_profit_pct"],
                "stop_loss":           trade["stop_loss"],
                "target_price":        trade["target_price"],
                "reward_risk_ratio":   trade["reward_risk_ratio"],
                "pattern_name":        trade.get("pattern_name", ""),
                "pattern_start":       trade.get("pattern_start"),
                "pattern_end":         trade.get("pattern_end"),
                "true_breakout_date":  trade.get("true_breakout_date"),
                "days_since_breakout": trade.get("days_since_breakout"),
                "pattern_breakout_price": trade.get("pattern_breakout_price"),
                "atr_at_entry":        trade.get("atr_at_entry"),
                "adx_at_entry":        trade.get("adx_at_entry"),
                "rsi_at_entry":        trade.get("rsi_at_entry"),
                "estimated_days_to_target": trade.get("estimated_days_to_target"),
                "estimated_target_date":    trade.get("estimated_target_date"),
            })
            trades.append(entry)
        else:
            entry["no_trade_reason"] = r.get("no_trade_reason", "")
            no_trades.append(entry)

    trades.sort(key=lambda x: x.get("gross_profit_pct", 0), reverse=True)

    return {
        "run_meta": {
            "cutoff_date": cutoff_date,
            "target_days": target_days,
            "run_date":    datetime.now().isoformat(),
            "elapsed_sec": round(elapsed, 1),
            "total":       len(results),
            "trades":      len(trades),
            "no_trades":   len(no_trades),
            "errors":      len(errors),
        },
        "trades":    trades,
        "no_trades": no_trades,
        "errors":    errors,
    }


def _build_markdown(summary: Dict[str, Any]) -> str:
    meta   = summary["run_meta"]
    trades = summary["trades"]
    no_tr  = summary["no_trades"]
    errs   = summary["errors"]

    lines = [
        f"# Halal Universe — Trade Predictions",
        f"",
        f"**Cutoff date**: {meta['cutoff_date']}  |  **Max window**: {meta['target_days']} trading days  |  "
        f"**Run**: {meta['run_date'][:16]}  |  **Elapsed**: {meta['elapsed_sec']}s",
        f"",
        f"## Summary",
        f"",
        f"| | Count |",
        f"|---|---|",
        f"| Total analyzed | {meta['total']} |",
        f"| **Active trades** (pattern-driven) | **{meta['trades']}** |",
        f"| No trade (stale / neutral / no pattern) | {meta['no_trades']} |",
        f"| Errors (data unavailable) | {meta['errors']} |",
        f"",
    ]

    # Sector breakdown of trades
    sector_trades: Dict[str, int] = {}
    sector_outcomes: Dict[str, Dict[str, int]] = {}
    for t in trades:
        s = t["sector"]
        sector_trades[s] = sector_trades.get(s, 0) + 1
        if s not in sector_outcomes:
            sector_outcomes[s] = {}
        o = t.get("exit_outcome", "?")
        sector_outcomes[s][o] = sector_outcomes[s].get(o, 0) + 1

    if sector_trades:
        lines += [
            "## Trades by Sector",
            "",
            "| Sector | Trades | OPEN | HIT_TARGET | HIT_STOP | EXPIRED |",
            "|--------|--------|------|------------|----------|---------|",
        ]
        for s in sorted(sector_trades):
            oc = sector_outcomes.get(s, {})
            lines.append(
                f"| {s} | {sector_trades[s]} | "
                f"{oc.get('OPEN',0)} | {oc.get('HIT_TARGET',0)} | {oc.get('HIT_STOP',0)} | {oc.get('EXPIRED',0)} |"
            )
        lines.append("")

    # Active trade table
    if trades:
        hit_t = [t for t in trades if t.get("exit_outcome") == "HIT_TARGET"]
        hit_s = [t for t in trades if t.get("exit_outcome") == "HIT_STOP"]
        exp   = [t for t in trades if t.get("exit_outcome") == "EXPIRED"]
        open_ = [t for t in trades if t.get("exit_outcome") == "OPEN"]

        def _trade_section(title, items):
            if not items:
                return []
            hdr = [
                f"## {title}",
                "",
                "| Ticker | Sector | Entry Date | Entry $ | Exit Date | Exit $ | "
                "Gross % | Net % | Pattern | R/R | Stop | Target |",
                "|--------|--------|------------|---------|-----------|--------|"
                "---------|-------|---------|-----|------|--------|",
            ]
            rows = []
            for t in items:
                rows.append(
                    f"| **{t['ticker']}** | {t['sector']} | {t['entry_date']} | "
                    f"{t['entry_price']:.2f} | {t['exit_date']} | {t['exit_price']:.2f} | "
                    f"{t.get('gross_profit_pct',0):+.2f}% | {t.get('net_profit_pct',0):+.2f}% | "
                    f"{t.get('entry_source','').replace('pattern:','')} | "
                    f"{t.get('reward_risk_ratio') or '-'} | "
                    f"{t.get('stop_loss',0):.2f} | {t.get('target_price',0):.2f} |"
                )
            return hdr + rows + [""]

        lines += _trade_section("OPEN — Live Setups (enter Monday 2026-04-06)", open_)
        lines += _trade_section("HIT_TARGET — Winning Trades (simulated)", hit_t)
        lines += _trade_section("EXPIRED — Held Full Window", exp)
        lines += _trade_section("HIT_STOP — Stopped Out", hit_s)

    # Error summary (top 10)
    if errs:
        lines += [
            f"## Errors ({len(errs)} tickers — insufficient data)",
            "",
            "| Ticker | Sector | Reason |",
            "|--------|--------|--------|",
        ]
        for e in errs[:30]:
            msg = str(e.get("error", ""))[:80]
            lines.append(f"| {e.get('ticker','?')} | {e.get('sector','?')} | {msg} |")
        if len(errs) > 30:
            lines.append(f"| ... | ... | +{len(errs)-30} more errors omitted |")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _export_excel(summary: Dict[str, Any], path: Path) -> None:
    """Export summary to a multi-sheet Excel workbook using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Helpers ──────────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TARGET_FILL = PatternFill("solid", fgColor="C6EFCE")
    STOP_FILL   = PatternFill("solid", fgColor="FFC7CE")
    OPEN_FILL   = PatternFill("solid", fgColor="FFEB9C")
    EXPIRED_FILL= PatternFill("solid", fgColor="DDEBF7")

    OUTCOME_FILL = {
        "HIT_TARGET": TARGET_FILL,
        "HIT_STOP":   STOP_FILL,
        "OPEN":       OPEN_FILL,
        "EXPIRED":    EXPIRED_FILL,
    }

    def _write_sheet(ws, headers: List[str], rows: List[List]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        # Auto-width
        for col_idx, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # ── Sheet 1: All Trades ───────────────────────────────────────────────
    ws_trades = wb.active
    ws_trades.title = "Trades"
    trade_headers = [
        "Ticker", "Sector", "Sentiment", "Confidence Score",
        "Tech Score", "Fund Score", "Conflict",
        "Entry Date", "Entry Price", "Entry Source",
        "Exit Date", "Exit Price", "Exit Outcome",
        "Holding Days", "Gross P&L %", "Net P&L %",
        "Stop Loss", "Target Price", "R/R Ratio",
        "No Trade Reason",
    ]
    trade_rows = []
    for t in summary["trades"] + summary["no_trades"]:
        trade_rows.append([
            t.get("ticker", ""),
            t.get("sector", ""),
            t.get("sentiment", ""),
            t.get("confidence_score", ""),
            t.get("tech_score", ""),
            t.get("fund_score", ""),
            t.get("conflict", ""),
            t.get("entry_date", ""),
            t.get("entry_price", ""),
            t.get("entry_source", ""),
            t.get("exit_date", ""),
            t.get("exit_price", ""),
            t.get("exit_outcome", "NO_TRADE"),
            t.get("holding_days", ""),
            t.get("gross_profit_pct", ""),
            t.get("net_profit_pct", ""),
            t.get("stop_loss", ""),
            t.get("target_price", ""),
            t.get("reward_risk_ratio", ""),
            t.get("no_trade_reason", ""),
        ])
    _write_sheet(ws_trades, trade_headers, trade_rows)

    # Colour rows by outcome
    for row_idx in range(2, ws_trades.max_row + 1):
        outcome = str(ws_trades.cell(row_idx, 13).value or "")
        fill = OUTCOME_FILL.get(outcome)
        if fill:
            for col_idx in range(1, len(trade_headers) + 1):
                ws_trades.cell(row_idx, col_idx).fill = fill

    # ── Sheet 2: Active Trades only (sorted by Net P&L desc) ─────────────
    ws_active = wb.create_sheet("Active Trades")
    active_rows = sorted(
        [r for r in trade_rows if r[12] != "NO_TRADE"],
        key=lambda r: (r[15] if isinstance(r[15], (int, float)) else -999),
        reverse=True,
    )
    _write_sheet(ws_active, trade_headers, active_rows)
    for row_idx in range(2, ws_active.max_row + 1):
        outcome = str(ws_active.cell(row_idx, 13).value or "")
        fill = OUTCOME_FILL.get(outcome)
        if fill:
            for col_idx in range(1, len(trade_headers) + 1):
                ws_active.cell(row_idx, col_idx).fill = fill

    # ── Sheet 3: Summary Stats ────────────────────────────────────────────
    ws_meta = wb.create_sheet("Summary")
    meta = summary["run_meta"]
    trades = summary["trades"]
    hit_t  = [t for t in trades if t.get("exit_outcome") == "HIT_TARGET"]
    hit_s  = [t for t in trades if t.get("exit_outcome") == "HIT_STOP"]
    exp    = [t for t in trades if t.get("exit_outcome") == "EXPIRED"]
    open_  = [t for t in trades if t.get("exit_outcome") == "OPEN"]

    def _avg(lst, key):
        vals = [v.get(key, 0) for v in lst if isinstance(v.get(key), (int, float))]
        return round(sum(vals) / len(vals), 2) if vals else ""

    stats = [
        ["Cutoff Date",         meta["cutoff_date"]],
        ["Target Days",         meta["target_days"]],
        ["Run Date",            meta["run_date"][:16]],
        ["Elapsed (s)",         meta["elapsed_sec"]],
        ["Total Tickers",       meta["total"]],
        ["Active Trades",       meta["trades"]],
        ["No Trade",            meta["no_trades"]],
        ["Errors",              meta["errors"]],
        [],
        ["Outcome",             "Count", "Avg Net P&L %"],
        ["HIT_TARGET",          len(hit_t), _avg(hit_t, "net_profit_pct")],
        ["HIT_STOP",            len(hit_s), _avg(hit_s, "net_profit_pct")],
        ["EXPIRED",             len(exp),   _avg(exp, "net_profit_pct")],
        ["OPEN",                len(open_), ""],
        [],
        ["Hit Rate (TARGET / all closed)",
         f"{len(hit_t) / max(len(hit_t)+len(hit_s)+len(exp), 1) * 100:.1f}%" if trades else "N/A"],
        ["Avg Net P&L (all trades)", _avg(trades, "net_profit_pct")],
    ]
    for row in stats:
        ws_meta.append(row)
    ws_meta.column_dimensions["A"].width = 38
    ws_meta.column_dimensions["B"].width = 18
    ws_meta.column_dimensions["C"].width = 18

    # ── Sheet 4: Errors ───────────────────────────────────────────────────
    ws_errors = wb.create_sheet("Errors")
    _write_sheet(
        ws_errors,
        ["Ticker", "Sector", "Error"],
        [[e.get("ticker",""), e.get("sector",""), str(e.get("error",""))[:200]]
         for e in summary["errors"]],
    )

    wb.save(path)
    print(f"✓  Excel saved to {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workers", type=int, default=5,
                        help="Parallel workers (default 5 — respects Polygon rate limits)")
    parser.add_argument("--date", type=str, default=None,
                        help="Cutoff date YYYY-MM-DD (default: last trading day)")
    parser.add_argument("--target-days", type=int, default=20)
    parser.add_argument("--sector", type=str, default=None,
                        help="Filter to one sector only")
    parser.add_argument("--universe", choices=["halal", "sectors"], default="halal",
                        help="'halal' = full Musaffa universe (default); 'sectors' = 50-ticker sector set")
    args = parser.parse_args()

    # Resolve cutoff
    if args.date:
        cutoff = datetime.strptime(args.date, "%Y-%m-%d").date()
    else:
        cutoff = _last_trading_day(date.today())
    cutoff_str = cutoff.isoformat()

    # Load tickers
    sector_map = _load_tickers(args.sector, universe=args.universe)
    flat: List[Tuple[str, str]] = []
    for sector, tickers in sector_map.items():
        for t in tickers:
            flat.append((t, sector))

    total = len(flat)
    print(f"\nBatch Prediction  [{args.universe.upper()} universe]")
    print(f"Cutoff: {cutoff_str}  |  Target days: {args.target_days}  |  Workers: {args.workers}")
    print(f"Tickers: {total} across {len(sector_map)} sectors")
    print("─" * 70)

    # Output directory
    tag = f"{args.universe}_{cutoff_str}"
    out_dir = OUTPUT_BASE / tag
    out_dir.mkdir(parents=True, exist_ok=True)

    results: List[Dict[str, Any]] = []
    start = time.time()

    with ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = {
            pool.submit(_predict_ticker, ticker, sector, cutoff_str, args.target_days): (ticker, sector)
            for ticker, sector in flat
        }
        done = 0
        for future in as_completed(futures):
            ticker, sector, result = future.result()
            done += 1
            if result:
                result["ticker"] = ticker
                result["sector"] = sector
                results.append(result)
                # Save individual file
                (out_dir / f"{ticker}.json").write_text(
                    json.dumps(result, indent=2, default=str)
                )
            eta = (time.time() - start) / done * (total - done)
            print(f"  Progress: {done}/{total}  ETA: {eta:.0f}s", end="\r", flush=True)

    elapsed = time.time() - start
    print(f"\n\nDone in {elapsed:.0f}s  ({elapsed/max(len(results),1):.1f}s/ticker avg)")

    # Build and save summary
    summary = _build_summary(results, cutoff_str, args.target_days, elapsed)
    md      = _build_markdown(summary)

    summary_json = OUTPUT_BASE / f"{tag}_summary.json"
    summary_md   = OUTPUT_BASE / f"{tag}_summary.md"
    excel_path   = OUTPUT_BASE / f"{tag}_results.xlsx"

    summary_json.write_text(json.dumps(summary, indent=2, default=str))
    summary_md.write_text(md)
    _export_excel(summary, excel_path)

    print(f"\nResults saved to:")
    print(f"  {out_dir}  ({len(results)} individual JSONs)")
    print(f"  {summary_json}")
    print(f"  {summary_md}")
    print(f"  {excel_path}  ← Excel")
    print(f"\n── Quick Stats ─────────────────────────────────────────")
    meta = summary["run_meta"]
    print(f"  Trades generated : {meta['trades']}")
    print(f"  No-trade         : {meta['no_trades']}")
    print(f"  Errors           : {meta['errors']}")

    trades = summary["trades"]
    if trades:
        hit_t = [t for t in trades if t.get("exit_outcome") == "HIT_TARGET"]
        hit_s = [t for t in trades if t.get("exit_outcome") == "HIT_STOP"]
        exp   = [t for t in trades if t.get("exit_outcome") == "EXPIRED"]
        open_ = [t for t in trades if t.get("exit_outcome") == "OPEN"]
        print(f"  OPEN (live)      : {len(open_)}")
        print(f"  HIT_TARGET       : {len(hit_t)}")
        print(f"  EXPIRED          : {len(exp)}")
        print(f"  HIT_STOP         : {len(hit_s)}")
        if exp:
            avg_pct = sum(t.get("net_profit_pct", 0) for t in exp) / len(exp)
            print(f"  Avg net % (expired): {avg_pct:+.2f}%")
        print("\nTop 10 by gross profit:")
        for t in trades[:10]:
            print(f"  {t['ticker']:<6} {t.get('exit_outcome','?'):<12} {t.get('gross_profit_pct',0):+6.2f}%  via {t.get('entry_source','')}")


if __name__ == "__main__":
    main()
